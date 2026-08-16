"""
Phase 3: MinerU OCR batch processing.

Takes all downloaded PDFs from the papers/ directory, sends them to MinerU
for text extraction, and produces a final Excel with two additional columns:
  - download_success (Yes/No)
  - ocr_success      (Yes/No/N/A)
"""

import csv
import json
import logging
import os
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Callable

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent

StatusCallback = Callable[[str, dict], None] | None

# Statuses that count as "successfully downloaded"
DOWNLOAD_OK = {"downloaded", "already_downloaded", "manual_added"}


def _find_latest_results_csv(output_dir: Path) -> Path | None:
    """Find the most recent results_final_*.csv in the output directory."""
    candidates = sorted(
        output_dir.glob("results_final_*.csv"),
        key=os.path.getmtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _read_results_csv(csv_path: str) -> list[dict]:
    """Read the pipeline results CSV into a list of dicts."""
    papers = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            papers.append(row)
    return papers


def _check_ocr_success(ocr_output_dir: Path, pdf_filename: str) -> bool:
    """
    Check whether MinerU successfully produced markdown for a given PDF.

    MinerU output layout:
      {ocr_output_dir}/{pdf_stem}/auto/{pdf_stem}.md
    """
    if not pdf_filename:
        return False
    stem = Path(pdf_filename).stem
    md_path = ocr_output_dir / stem / "auto" / f"{stem}.md"
    if not md_path.exists():
        return False
    # File exists — check it has actual content (not empty)
    try:
        return md_path.stat().st_size > 50  # arbitrary minimum: real content
    except OSError:
        return False


def _pre_scan_corrupt_pdfs(
    papers_dir: Path,
    quarantine_dir: Path,
) -> list[str]:
    """
    Scan papers/ for truncated/corrupt PDFs before invoking MinerU.

    A truncated download (e.g. cut off at a fixed 256 KB buffer) has a valid
    ``%PDF`` header but is missing the ``%%EOF`` trailer.  pypdfium2 (used by
    MinerU's ``collect_input_documents`` probe) fails on such files with
    "Data format error", which crashes the ENTIRE batch with exit code 1 --
    even though every other PDF is fine.

    We detect these cheaply (no extra dependencies) by checking that the
    trailing bytes contain ``%%EOF``, then move the bad files into a
    quarantine directory so MinerU never sees them.

    Returns the list of quarantined filenames.
    """
    quarantine_dir.mkdir(parents=True, exist_ok=True)
    quarantined: list[str] = []

    for pdf in sorted(papers_dir.glob("*.pdf")):
        try:
            with open(pdf, "rb") as fh:
                fh.seek(0, 2)
                size = fh.tell()
                fh.seek(max(0, size - 2048))
                tail = fh.read()
            if b"%%EOF" in tail:
                continue
        except OSError:
            pass  # unreadable file → treat as corrupt

        # Corrupt: missing %%EOF trailer (or unreadable)
        try:
            shutil.move(str(pdf), str(quarantine_dir / pdf.name))
            quarantined.append(pdf.name)
            logger.warning(f"Phase 3: quarantined corrupt PDF: {pdf.name}")
        except OSError as e:
            logger.error(f"Phase 3: failed to quarantine {pdf.name}: {e}")

    return quarantined


def _write_final_excel(
    papers: list[dict],
    ocr_output_dir: Path,
    excel_path: Path,
) -> dict:
    """
    Write the final Excel file with download_success and ocr_success columns.

    Returns a stats dict.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # ── Header styling ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
    green_fill = PatternFill(start_color="E1F5EE", end_color="E1F5EE", fill_type="solid")
    red_fill = PatternFill(start_color="FAECE7", end_color="FAECE7", fill_type="solid")
    gray_fill = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid")

    # ── Headers ──
    headers = [
        "DOI", "Title", "First Author", "Year", "Journal",
        "Filename", "Download Status", "Phase", "Filepath", "Message",
        "download_success", "ocr_success",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Data rows ──
    total = len(papers)
    dl_ok = 0
    ocr_ok = 0
    ocr_attempted = 0

    for paper in papers:
        status = paper.get("status", "")
        filename = paper.get("filename", "")
        dl_success = status in DOWNLOAD_OK
        if dl_success:
            dl_ok += 1

        if dl_success and filename:
            ocr_attempted += 1
            ocr_success = _check_ocr_success(ocr_output_dir, filename)
            if ocr_success:
                ocr_ok += 1
        else:
            ocr_success = False

        row = [
            paper.get("doi", ""),
            paper.get("title", ""),
            paper.get("first_author", ""),
            paper.get("year", ""),
            paper.get("journal", ""),
            filename,
            status,
            paper.get("phase", ""),
            paper.get("filepath", ""),
            paper.get("message", ""),
            "Yes" if dl_success else "No",
            "Yes" if ocr_success else ("N/A" if not dl_success else "No"),
        ]
        ws.append(row)

        # Row coloring
        row_idx = ws.max_row
        ocr_cell = ws.cell(row=row_idx, column=12)
        dl_cell = ws.cell(row=row_idx, column=11)
        if ocr_cell.value == "Yes":
            ocr_cell.fill = green_fill
        elif ocr_cell.value == "No":
            ocr_cell.fill = red_fill
        else:
            ocr_cell.fill = gray_fill
        if dl_cell.value == "Yes":
            dl_cell.fill = green_fill
        else:
            dl_cell.fill = red_fill

    # ── Column widths ──
    col_widths = {
        "A": 28, "B": 40, "C": 14, "D": 8, "E": 20,
        "F": 30, "G": 16, "H": 8, "I": 40, "J": 30,
        "K": 16, "L": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = ws.dimensions

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))

    return {
        "total_papers": total,
        "downloaded": dl_ok,
        "ocr_attempted": ocr_attempted,
        "ocr_success": ocr_ok,
        "ocr_failed": ocr_attempted - ocr_ok,
        "not_downloaded": total - dl_ok,
        "excel_path": str(excel_path),
    }


def run_phase3(
    results_csv_path: str = "",
    output_dir: str = "",
    mineru_exe: str = "",
    mineru_backend: str = "pipeline",
    ocr_subdir: str = "ocr_output",
    ocr_output_dir: str = "",
    progress_callback: StatusCallback = None,
) -> dict:
    """
    Run Phase 3: MinerU batch OCR on all downloaded PDFs.

    MinerU is called as a standalone CLI subprocess — no API/Hermes bridging.

    Args:
        results_csv_path: Path to the results_final_*.csv from Phase 1+2.
                          If empty, auto-detect the latest one in output_dir.
        output_dir: Root output directory (contains papers/).
        mineru_exe: Path to the MinerU CLI executable.
        mineru_backend: MinerU backend ("pipeline" or "vlm-engine").
        ocr_subdir: Fallback subdirectory name for MinerU output (under output_dir).
                    Only used when ocr_output_dir is not provided.
        ocr_output_dir: Explicit path for MinerU MD output. Takes priority over ocr_subdir.
                        If empty, falls back to {output_dir}/{ocr_subdir}.
        progress_callback: Optional callback(status, data) for real-time updates.

    Returns:
        Stats dict with counts and paths.
    """
    # ── Resolve paths ──
    out_root = Path(output_dir).resolve() if output_dir else BASE_DIR / "output"
    papers_dir = out_root / "papers"

    # MinerU output directory: explicit path > subdir under output_dir
    if ocr_output_dir:
        ocr_output_dir = Path(ocr_output_dir).resolve()
    else:
        ocr_output_dir = out_root / ocr_subdir
    ocr_output_dir.mkdir(parents=True, exist_ok=True)

    # ── Find results CSV ──
    if not results_csv_path:
        found = _find_latest_results_csv(out_root)
        if not found:
            msg = "No results_final_*.csv found. Run Phase 1+2 first."
            logger.error(msg)
            if progress_callback:
                progress_callback("phase3_error", {"error": msg})
            return {"error": msg}
        results_csv_path = str(found)

    logger.info(f"Phase 3: reading results from {results_csv_path}")
    if progress_callback:
        progress_callback("phase3_starting", {
            "csv": results_csv_path,
            "papers_dir": str(papers_dir),
            "ocr_output": str(ocr_output_dir),
        })

    papers = _read_results_csv(results_csv_path)
    if not papers:
        msg = "Results CSV is empty. Nothing to OCR."
        if progress_callback:
            progress_callback("phase3_error", {"error": msg})
        return {"error": msg}

    # ── Pre-scan: quarantine truncated/corrupt PDFs before MinerU ──
    # A single corrupt PDF (valid %PDF header but missing %%EOF trailer)
    # crashes MinerU's whole batch with exit code 1.  Move them out first.
    quarantine_dir = papers_dir.parent / "papers_corrupt"
    quarantined = _pre_scan_corrupt_pdfs(papers_dir, quarantine_dir)
    if quarantined:
        logger.warning(
            f"Phase 3: quarantined {len(quarantined)} corrupt PDF(s) → {quarantine_dir}"
        )
        if progress_callback:
            progress_callback("phase3_warning", {
                "message": f"隔离了 {len(quarantined)} 个损坏 PDF（见 papers_corrupt/）",
            })

    # ── Count actual PDFs to process ──
    pdf_files = list(papers_dir.glob("*.pdf"))
    download_ok_papers = [p for p in papers if p.get("status", "") in DOWNLOAD_OK]

    logger.info(
        f"Phase 3: {len(download_ok_papers)} downloaded papers, "
        f"{len(pdf_files)} PDF files found on disk"
    )
    if progress_callback:
        progress_callback("phase3_preparing", {
            "total_papers": len(papers),
            "downloaded": len(download_ok_papers),
            "pdfs_on_disk": len(pdf_files),
        })

    if not pdf_files:
        msg = "No PDF files found in papers/ directory. Nothing to OCR."
        if progress_callback:
            progress_callback("phase3_error", {"error": msg})
        return {"error": msg}

    # ── Check MinerU executable ──
    if not mineru_exe or not Path(mineru_exe).exists():
        msg = (
            f"MinerU executable not found at: {mineru_exe}\n"
            "Please check MINERU_EXECUTABLE in config.py."
        )
        logger.error(msg)
        if progress_callback:
            progress_callback("phase3_error", {"error": msg})
        return {"error": msg}

    # ── Invoke MinerU CLI ──
    cmd = [
        mineru_exe,
        "-p", str(papers_dir),
        "-o", str(ocr_output_dir),
        "-b", mineru_backend,
    ]

    logger.info(f"Phase 3: running MinerU: {' '.join(cmd)}")
    if progress_callback:
        progress_callback("phase3_running", {
            "command": " ".join(cmd),
            "total_pdfs": len(pdf_files),
        })

    try:
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            cwd=str(Path(mineru_exe).parent.parent),  # MinerU project root
        )

        # Stream output line by line
        line_count = 0
        for line in process.stdout:
            line = line.rstrip()
            if not line:
                continue
            line_count += 1
            logger.info(f"[MinerU] {line}")

            # Forward every line via WebSocket for full transparency
            if progress_callback:
                progress_callback("phase3_log", {
                    "line": line[:200],
                    "line_count": line_count,
                })

        process.wait()
        return_code = process.returncode

        if return_code != 0:
            logger.warning(f"MinerU exited with code {return_code} (output may still be valid)")
            if progress_callback:
                progress_callback("phase3_warning", {
                    "message": f"MinerU exited with code {return_code} (output may still be valid)",
                })
        else:
            logger.info("MinerU completed successfully")
            if progress_callback:
                progress_callback("phase3_done", {
                    "message": "MinerU processing complete",
                })

    except Exception as e:
        msg = f"Failed to run MinerU: {e}"
        logger.error(msg)
        if progress_callback:
            progress_callback("phase3_error", {"error": msg})
        return {"error": msg}

    # ── Scan OCR output and generate Excel ──
    if progress_callback:
        progress_callback("phase3_generating_excel", {
            "ocr_output_dir": str(ocr_output_dir),
        })

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = ocr_output_dir / f"results_final_with_ocr_{timestamp}.xlsx"

    stats = _write_final_excel(papers, ocr_output_dir, excel_path)

    logger.info(
        f"Phase 3 complete: {stats['ocr_success']}/{stats['ocr_attempted']} OCR success, "
        f"Excel: {excel_path}"
    )
    if progress_callback:
        progress_callback("phase3_completed", stats)

    return stats
