"""
FastAPI backend for DoiHarvest unified pipeline.

Provides:
- POST /api/upload  — upload CSV
- POST /api/run     — start pipeline
- WS  /ws/progress  — real-time status updates
- GET  /api/state   — current pipeline state
- GET  /api/download/<filename> — download results
"""

import asyncio
import csv
import json
import logging
import os
import re
import shutil
import threading
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, Form, UploadFile, File, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from .pipeline import run_phase1_pipeline, run_phase2_pipeline, _load_state, _save_state
from .phase3_mineru import run_phase3
from .phase4_screening import run_phase4

BASE_DIR = Path(__file__).resolve().parent.parent

# Load user configuration
import importlib.util
_config_path = BASE_DIR / "config.py"
_cfg = None
if _config_path.exists():
    spec = importlib.util.spec_from_file_location("config", _config_path)
    _cfg = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(_cfg)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger("api")

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
FRONTEND_DIR = BASE_DIR / "frontend"

DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Track the active input CSV so the pipeline always uses the most recent user action
_active_input_csv: str | None = None

app = FastAPI(title="DoiHarvest", version="2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# WebSocket connections
active_connections: list[WebSocket] = []
pipeline_thread: threading.Thread | None = None
pipeline_running = False
phase2_running = False
phase3_running = False
phase4_running = False

# Captured main event loop for cross-thread WebSocket broadcasting
_main_loop: asyncio.AbstractEventLoop | None = None


@app.on_event("startup")
async def _capture_event_loop():
    """Capture the main asyncio event loop so background threads can use it."""
    global _main_loop
    _main_loop = asyncio.get_running_loop()
    logger.info("Main event loop captured for cross-thread WebSocket broadcast")


async def broadcast(message: dict) -> None:
    """Send message to all connected WebSocket clients."""
    disconnected = []
    for ws in active_connections:
        try:
            await ws.send_json(message)
        except Exception:
            disconnected.append(ws)
    for ws in disconnected:
        active_connections.remove(ws)


def _load_state_healed() -> dict:
    """Load pipeline state and self-heal stale 'running' statuses.

    After a server restart, a status left as 'phase1_running' or
    'phase2_running' (the process was killed mid-run) is stale. If the
    corresponding in-process flag is False, correct it so the UI doesn't
    get stuck thinking a phase is still in progress and never advance.
    """
    state = _load_state()
    status = state.get("status")
    if status == "phase2_running" and not phase2_running:
        state["status"] = "awaiting_phase2"
        _save_state(state)
    elif status == "phase1_running" and not pipeline_running:
        state["status"] = ("awaiting_phase2"
                           if state.get("phase1_complete") else "starting")
        _save_state(state)
    return state


def progress_callback(status: str, data: dict) -> None:
    """Thread-safe callback to broadcast progress via WebSocket.

    Uses the pre-captured _main_loop to schedule the async broadcast
    from a background thread (pipeline runs in a daemon thread).
    """
    if _main_loop is None or not _main_loop.is_running():
        return
    message = {"type": status, "data": data, "state": _load_state()}
    asyncio.run_coroutine_threadsafe(broadcast(message), _main_loop)


# ── Workspace config ────────────────────────────────
CONFIG_FILE = BASE_DIR / "workspace_config.json"

def _load_workspace_config() -> dict:
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def _save_workspace_config(cfg: dict) -> None:
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def _get_output_dir() -> Path:
    """Resolve the current output directory from workspace config."""
    cfg = _load_workspace_config()
    path_str = cfg.get("output_dir", "").strip()
    if path_str:
        path = Path(path_str)
        if not path.is_absolute():
            path = BASE_DIR / path
        return path.resolve()
    return BASE_DIR / "output"


def _get_mineru_output_dir() -> Path:
    """Resolve the MinerU OCR output directory (where .md and final Excel go)."""
    cfg = _load_workspace_config()
    explicit = cfg.get("mineru_output_dir", "").strip()
    if explicit:
        path = Path(explicit)
        if not path.is_absolute():
            path = BASE_DIR / path
        return path.resolve()
    # Fallback to {output_dir}/ocr_output
    return _get_output_dir() / "ocr_output"


# ── Results CSV helpers ──────────────────────────────

def _read_results_csv(filepath: str | Path) -> list[dict]:
    """Read results CSV into list of dicts."""
    results = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            results.append(dict(row))
    return results


def _find_latest_results_csv(out_root: Path) -> Path | None:
    """Find the most recent results_final_*.csv in the output directory."""
    candidates = sorted(
        out_root.glob("results_final_*.csv"),
        key=os.path.getmtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _sanitize_filename(name: str) -> str:
    """Remove unsafe characters from filename."""
    return re.sub(r'[<>:"/\\|?*]', '_', str(name))


def _build_paper_filename(paper: dict, papers_dir: Path | None = None) -> str:
    """Build an 'Author_Year_Title.pdf' filename from a CSV row's metadata.

    Falls back to 'manual_{doi}.pdf' if metadata is incomplete.
    If papers_dir is provided and the generated name already exists,
    appends a numeric suffix to avoid collisions.
    """
    import unicodedata

    def _clean(s: str, max_len: int = 150) -> str:
        s = unicodedata.normalize("NFKC", str(s or ""))
        s = s.replace("\ufeff", "").replace("\u200b", "").replace("\u00a0", " ")
        s = re.sub(r'[<>:"/\\|?*]', "_", s)
        s = re.sub(r"\s+", " ", s).strip(" _.")
        s = re.sub(r"_+", "_", s)
        if len(s) > max_len:
            s = s[:max_len].rsplit("_", 1)[0]
        return s

    author = _clean(paper.get("first_author", ""), 60)
    year = _clean(paper.get("year", ""), 8)
    title = _clean(paper.get("title", ""), 120)

    if not author and not title:
        # Fallback to DOI-based name
        doi = paper.get("doi", "").strip().lower().replace("/", "_")
        return f"manual_{_sanitize_filename(doi)}.pdf"

    if not author:
        author = "UnknownAuthor"
    if not year:
        year = "0000"
    if not title:
        title = "Untitled"

    base = f"{author}_{year}_{title}.pdf"
    if papers_dir is None:
        return base

    # Avoid collision
    candidate = base
    counter = 2
    while (papers_dir / candidate).exists():
        stem = base.rsplit(".", 1)[0]
        candidate = f"{stem}_{counter}.pdf"
        counter += 1
    return candidate


def _update_paper_status(out_root: Path, doi: str, new_status: str, filepath: str = "") -> bool:
    """Update a paper's status in the latest results CSV when a PDF is manually added.
    Returns True if the paper was found and updated."""
    csv_path = _find_latest_results_csv(out_root)
    if not csv_path:
        return False

    papers = _read_results_csv(csv_path)
    doi_lower = doi.strip().lower()
    updated = False

    for p in papers:
        if p.get("doi", "").strip().lower() == doi_lower:
            p["status"] = new_status
            if filepath:
                p["filepath"] = filepath
                p["filename"] = Path(filepath).name
            updated = True
            break

    if not updated:
        # DOI not in results yet — add it as a new row
        papers.append({
            "doi": doi.strip(),
            "title": "",
            "first_author": "",
            "year": "",
            "journal": "",
            "filename": Path(filepath).name if filepath else "",
            "status": new_status,
            "phase": "0",
            "filepath": filepath,
            "message": "Manually added",
        })

    # Re-write CSV
    fieldnames = ["doi", "title", "first_author", "year", "journal",
                  "filename", "status", "phase", "filepath", "message"]
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(papers)

    logger.info(f"Updated paper status: {doi.strip()} → '{new_status}' in {csv_path.name}")
    return True


@app.get("/api/health")
async def health():
    return {"status": "ok"}


@app.get("/api/state")
async def get_state():
    return _load_state_healed()


@app.get("/api/config")
async def get_config():
    """Get current workspace configuration."""
    cfg = _load_workspace_config()
    if not cfg.get("output_dir"):
        cfg["output_dir"] = str(BASE_DIR / "output")
    if not cfg.get("mineru_output_dir"):
        cfg["mineru_output_dir"] = ""
    if not cfg.get("school_name"):
        cfg["school_name"] = "北京大学医学部"
    # Backward compat: show old keys as empty so frontend doesn't break
    if not cfg.get("portal_url"):
        cfg["portal_url"] = ""
    if not cfg.get("webvpn_base"):
        cfg["webvpn_base"] = ""
    return cfg


@app.post("/api/config")
async def save_config(data: dict):
    """Save workspace configuration (output_dir, mineru_output_dir, etc.)."""
    cfg = _load_workspace_config()

    if "output_dir" in data:
        new_dir = data["output_dir"].strip()
        if new_dir:
            path = Path(new_dir)
            if not path.is_absolute():
                path = (BASE_DIR / path).resolve()
            else:
                path = path.resolve()
            path.mkdir(parents=True, exist_ok=True)
            cfg["output_dir"] = str(path)
        else:
            return {"error": "output_dir cannot be empty"}

    if "mineru_output_dir" in data:
        new_dir = data["mineru_output_dir"].strip()
        if new_dir:
            path = Path(new_dir)
            if not path.is_absolute():
                path = (BASE_DIR / path).resolve()
            else:
                path = path.resolve()
            path.mkdir(parents=True, exist_ok=True)
            cfg["mineru_output_dir"] = str(path)
        else:
            # Empty string = clear it, fall back to default ocr_output subdir
            cfg["mineru_output_dir"] = ""

    if "school_name" in data:
        cfg["school_name"] = data["school_name"].strip()

    # DeepSeek / Phase 4 settings
    for key in ("deepseek_api_key", "deepseek_model", "deepseek_api_base"):
        if key in data:
            cfg[key] = data[key].strip()
    if "deepseek_max_content_chars" in data:
        try:
            cfg["deepseek_max_content_chars"] = int(data["deepseek_max_content_chars"])
        except (ValueError, TypeError):
            pass
    if "deepseek_temperature" in data:
        try:
            cfg["deepseek_temperature"] = float(data["deepseek_temperature"])
        except (ValueError, TypeError):
            pass
    # Screening criteria persistence
    if "screening_inclusion" in data:
        cfg["screening_inclusion"] = data["screening_inclusion"]
    if "screening_exclusion" in data:
        cfg["screening_exclusion"] = data["screening_exclusion"]

    # Backward compat: accept old fields, store silently
    if "portal_url" in data:
        cfg["portal_url"] = data["portal_url"].strip()
    if "webvpn_base" in data:
        cfg["webvpn_base"] = data["webvpn_base"].strip()
    if "library_url" in data:
        cfg["portal_url"] = data["library_url"].strip()

    _save_workspace_config(cfg)
    logger.info(f"Workspace config updated: {cfg}")
    return {"status": "ok", **cfg}


@app.post("/api/upload")
async def upload_csv(file: UploadFile = File(...)):
    """Upload a CSV file with DOI list."""
    if not file.filename or not file.filename.endswith(".csv"):
        return {"error": "Please upload a .csv file"}

    filepath = DATA_DIR / file.filename
    try:
        content = await file.read()
    except Exception as e:
        logger.exception("Failed to read uploaded file")
        return {"error": f"Failed to read file: {e}"}

    with open(filepath, "wb") as f:
        f.write(content)

    # Mark this file as the active input for the pipeline
    global _active_input_csv
    _active_input_csv = str(filepath)

    # New CSV upload starts a fresh workflow — clear stale Phase 1/2/3/4
    # completion so the UI doesn't block re-running and the WS state
    # restore (on reconnect) doesn't keep the Start button disabled.
    _st = _load_state()
    if _st.get("phase1_complete") or _st.get("csv_path"):
        for _k in ("phase1_complete", "phase1_stats", "phase2_ready",
                   "failed_count", "phase2_complete", "phase2_stats",
                   "summary", "phase3_complete", "phase4_complete"):
            _st.pop(_k, None)
        _st["status"] = "starting"
        _st["csv_path"] = str(filepath)
        _save_state(_st)

    # Quick validation: check for DOI column and count missing DOIs
    import csv, io
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("gbk")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))

    has_doi = False
    has_title = False
    doi_col = None
    title_col = None
    for fn in (reader.fieldnames or []):
        if "doi" in (fn or "").lower():
            has_doi = True
            doi_col = fn
        if "title" in (fn or "").lower():
            has_title = True
            title_col = fn

    row_count = 0
    missing_doi = 0
    for row in reader:
        row_count += 1
        doi_val = (row.get(doi_col, "") if doi_col else "").strip()
        if not doi_val:
            missing_doi += 1

    return {
        "filename": file.filename,
        "path": str(filepath),
        "rows": row_count,
        "has_doi_column": has_doi,
        "has_title_column": has_title,
        "missing_doi": missing_doi,
    }


# ── DOI Completion via Title Search ─────────────────

@app.post("/api/doi-completion")
def run_doi_completion():
    """
    Scan the latest uploaded CSV for rows with a Title but no DOI,
    then search Crossref by title to find candidate DOIs.

    Returns a structured list with confidence scores so the user can
    review and confirm each match in the frontend.

    NOTE: This is a regular (non-async) endpoint so FastAPI runs it in a
    thread pool.  That way ``progress_callback`` can use
    ``run_coroutine_threadsafe`` to push WebSocket progress messages
    while the search is running.
    """
    # Pick the active input CSV
    global _active_input_csv
    active = _active_input_csv
    if active and Path(active).exists():
        csv_path = Path(active)
    else:
        csv_files = sorted(DATA_DIR.glob("*.csv"), key=os.path.getmtime, reverse=True)
        if not csv_files:
            return {"error": "No CSV file found. Upload one first."}
        csv_path = csv_files[0]

    logger.info(f"DOI completion: scanning {csv_path}")

    # Read CSV, identify columns
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []

        # Detect DOI column
        doi_col = None
        for name in fieldnames:
            if "doi" in name.lower():
                doi_col = name
                break

        # Detect Title column
        title_col = None
        for name in fieldnames:
            if "title" in name.lower():
                title_col = name
                break

        # Detect Year column (optional)
        year_col = None
        for name in fieldnames:
            if "year" in name.lower() and "year" != name.lower():
                year_col = name
                break

        rows = list(reader)

    if not title_col:
        return {"error": "No 'Title' column found in CSV. Cannot perform DOI completion."}

    # Build papers list for those missing DOIs
    papers_missing_doi = []
    all_papers = []
    for idx, row in enumerate(rows):
        doi = (row.get(doi_col, "") if doi_col else "").strip()
        title = row.get(title_col, "").strip()
        year = row.get(year_col, "").strip() if year_col else ""

        all_papers.append({
            "doi": doi,
            "title": title,
            "year": year,
            "row": dict(row),
        })

        if not doi and title:
            papers_missing_doi.append({
                "row_index": idx,
                "title": title,
                "year": year,
            })

    if not papers_missing_doi:
        return {
            "total_rows": len(rows),
            "missing_doi": 0,
            "has_doi": sum(1 for p in all_papers if p["doi"]),
            "message": "All rows already have DOIs. No completion needed.",
            "results": [],
            "stats": {},
        }

    logger.info(
        f"DOI completion: {len(papers_missing_doi)} papers without DOI, "
        f"{sum(1 for p in all_papers if p['doi'])} with DOI"
    )

    progress_callback("doi_completion_start", {
        "message": f"Starting DOI completion: {len(papers_missing_doi)} papers need DOIs",
        "total": len(papers_missing_doi),
    })

    from .metadata import batch_find_dois

    completion = batch_find_dois(
        papers_missing_doi,
        progress_callback=progress_callback,
    )

    stats = completion.get("stats", {})
    progress_callback("doi_completion_finish", {
        "message": (
            f"DOI completion finished: {stats.get('matched_high', 0)} high, "
            f"{stats.get('matched_medium', 0)} medium, "
            f"{stats.get('matched_low', 0)} low, "
            f"{stats.get('no_match', 0)} no match"
        ),
        "stats": stats,
    })

    # Determine which results the user should review (medium/low confidence)
    needs_review = sum(
        1 for r in completion["results"]
        if r["candidates"] and r["candidates"][0]["confidence"] != "high"
    )

    return {
        "total_rows": len(rows),
        "missing_doi": len(papers_missing_doi),
        "has_doi": sum(1 for p in all_papers if p["doi"]),
        "needs_review": needs_review,
        "csv_path": str(csv_path),
        "csv_name": csv_path.name,
        **completion,
    }


@app.post("/api/doi-completion/apply")
async def apply_doi_completion(data: dict):
    """
    Write user-confirmed DOIs back to a new CSV.

    Expects JSON: {
        "resolved": [
            {"row_index": 0, "doi": "10.xxx/yyy"},
            ...
        ],
        "ignored": [3, 7, 12],       # row indices to exclude from output
        "skip_empty": true             # if true, remove rows that still have no DOI
    }

    Ignored rows are kept in the output CSV (with empty DOI) so they flow into
    Phase 1/2 and are surfaced as "no_doi" for manual PDF supplementation before Phase 3.
    """
    resolved_map: dict[int, str] = {}
    for item in data.get("resolved", []):
        idx = item.get("row_index")
        doi = item.get("doi", "").strip()
        if idx is not None and doi:
            resolved_map[idx] = doi

    ignored_rows: set[int] = set(data.get("ignored", []))
    skip_empty = data.get("skip_empty", True)

    # Pick the active input CSV
    global _active_input_csv
    active = _active_input_csv
    if active and Path(active).exists():
        csv_path = Path(active)
    else:
        csv_files = sorted(DATA_DIR.glob("*.csv"), key=os.path.getmtime, reverse=True)
        if not csv_files:
            return {"error": "No CSV file found."}
        csv_path = csv_files[0]

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)

    # Detect DOI and Title columns
    doi_col = None
    title_col = None
    for name in fieldnames:
        if "doi" in name.lower():
            doi_col = name
        if "title" in name.lower():
            title_col = name

    if not doi_col:
        # No DOI column exists — insert one right after the first column (or Title column)
        if title_col and title_col in fieldnames:
            insert_pos = fieldnames.index(title_col) + 1
        else:
            insert_pos = 0
        doi_col = "DOI"
        fieldnames.insert(insert_pos, doi_col)
        for row in rows:
            row[doi_col] = ""
    elif title_col and doi_col not in fieldnames:
        # Detect case-insensitively
        for name in fieldnames:
            if name.lower() == "doi":
                doi_col = name
                break
        else:
            doi_col = "DOI"
            fieldnames.insert(0, doi_col)
            for row in rows:
                row[doi_col] = ""

    updated = 0
    skipped = 0
    ignored_count = len(ignored_rows)
    for idx, row in enumerate(rows):
        if idx in resolved_map:
            row[doi_col] = resolved_map[idx]
            updated += 1
        elif skip_empty and not row.get(doi_col, "").strip():
            skipped += 1

    # Build output rows: include resolved, keep existing DOIs, keep ignored rows too
    # (ignored no-DOI papers stay so they can flow into Phase 1/2 with "no_doi" status)
    new_rows = []
    for i, r in enumerate(rows):
        if i in resolved_map:
            new_rows.append(r)
        elif skip_empty and not r.get(doi_col, "").strip():
            continue  # skip_empty mode: drop rows that still have no DOI
        else:
            new_rows.append(r)

    # Write output CSV (append _completed suffix)
    stem = csv_path.stem
    if stem.endswith("_completed"):
        out_name = f"{stem}.csv"
    else:
        out_name = f"{stem}_completed.csv"
    out_path = csv_path.parent / out_name

    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(new_rows)

    logger.info(
        f"DOI completion applied: {updated} updated, "
        f"{ignored_count} kept (ignored no-DOI), {skipped} skipped (no DOI), "
        f"wrote {len(new_rows)} rows to {out_name}"
    )

    # Update active input to the completed file
    _active_input_csv = str(out_path)

    # New completed CSV → fresh workflow; clear stale phase completion
    _st = _load_state()
    if _st.get("phase1_complete") or _st.get("csv_path"):
        for _k in ("phase1_complete", "phase1_stats", "phase2_ready",
                   "failed_count", "phase2_complete", "phase2_stats",
                   "summary", "phase3_complete", "phase4_complete"):
            _st.pop(_k, None)
        _st["status"] = "starting"
        _st["csv_path"] = str(out_path)
        _save_state(_st)

    return {
        "status": "ok",
        "updated": updated,
        "ignored": ignored_count,
        "skipped": skipped,
        "output_rows": len(new_rows),
        "output_file": str(out_path),
        "output_name": out_name,
    }


@app.post("/api/run")
async def start_pipeline(data: dict | None = None):
    """Start Phase 1 (Sci-Hub download) on all DOIs from the CSV."""
    global pipeline_thread, pipeline_running, _active_input_csv

    if pipeline_running:
        return {"error": "Phase 1 (Sci-Hub) already running"}

    csv_path = _active_input_csv
    if not csv_path or not Path(csv_path).exists():
        csv_files = sorted(DATA_DIR.glob("*.csv"), key=os.path.getmtime, reverse=True)
        completed = [f for f in csv_files if f.stem.endswith("_completed")]
        if completed:
            csv_path = str(completed[0])
        elif csv_files:
            csv_path = str(csv_files[0])
        else:
            return {"error": "No CSV file found. Upload one first."}

    # Guard: if Phase 1 already completed for THIS csv, don't let the user
    # accidentally re-run the whole batch — route them to Phase 2 instead.
    # A different CSV (fresh upload) is allowed through. Pass {"force": true}
    # to override.
    state = _load_state()
    if state.get("phase1_complete") and not (data or {}).get("force"):
        p1_file = Path(_get_output_dir()) / "phase1_results.json"
        prior_csv = state.get("csv_path", "")
        if p1_file.exists() and prior_csv:
            try:
                same = (os.path.normpath(str(csv_path)).lower()
                        == os.path.normpath(prior_csv).lower())
            except Exception:
                same = False
            if same:
                p1s = state.get("phase1_stats", {})
                failed = state.get("failed_count", 0)
                return {"error": (
                    f"Phase 1 已完成（{p1s.get('downloaded', '?')} 下载 / "
                    f"{state.get('total_papers', '?')} 篇），无需重跑。"
                    f"{failed} 篇待 Phase 2 重试，请点击「Step 2: Execute Phase 2」。"
                )}

    logger.info(f"Starting Phase 1 (Sci-Hub) with {csv_path}")

    output_dir = str(_get_output_dir())
    pipeline_running = True

    def _run():
        global pipeline_running
        try:
            result = run_phase1_pipeline(
                csv_path=csv_path,
                unpaywall_email=getattr(_cfg, "UNPAYWALL_EMAIL", "") if _cfg else "",
                enable_unpaywall=getattr(_cfg, "ENABLE_UNPAYWALL", True) if _cfg else True,
                output_dir=output_dir,
                progress_callback=progress_callback,
            )
            logger.info(f"Phase 1 (Sci-Hub) finished: {result}")
        except Exception as e:
            logger.error(f"Phase 1 failed: {e}")
            progress_callback("error", {"error": str(e)})
        finally:
            pipeline_running = False

    pipeline_thread = threading.Thread(target=_run, daemon=True)
    pipeline_thread.start()

    return {"status": "started", "csv": csv_path, "phase": 1}


# ── Phase 2: Library WebVPN (PKU) ────────────────────

@app.post("/api/phase2/configure")
async def configure_webvpn(data: dict):
    """Configure WebVPN cookies for the PKU library proxy.

    Expects JSON: {"provider": "pku", "cookies": {"key1": "value1", ...}}

    How to get cookies:
      1. Open https://wpn.pku.edu.cn in your browser
      2. Log in with PKU credentials
      3. Open DevTools → Application → Cookies
      4. Copy all cookie name=value pairs
    """
    from .phase2_webvpn import configure_webvpn as _configure

    cookies = data.get("cookies", {})
    base_url = data.get("base_url", "")
    provider = data.get("provider", "")

    try:
        cfg = _configure(cookies=cookies, base_url=base_url, provider=provider)
        active = cfg.get("provider", "pku")
        has_cookies = bool(cfg.get("cookies"))
        return {
            "status": "configured" if has_cookies else "no_cookies",
            "provider": active,
            "cookie_count": len(cfg.get("cookies", {})),
            "base_url": cfg.get("base_url", ""),
            "message": (
                f"{active} WebVPN configured with {len(cfg.get('cookies', {}))} cookies."
                if has_cookies
                else "No cookies provided. Paste cookies from your browser."
            ),
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/phase2/config")
async def get_webvpn_config():
    """Get WebVPN configuration status for all providers."""
    from .phase2_webvpn import load_webvpn_config, PROVIDERS

    cfg = load_webvpn_config()
    providers_status = {}
    for pid, prov in PROVIDERS.items():
        sect = cfg.get(pid, {})
        # direct mode (campus VPN client) needs no cookies — always "configured"
        is_direct = prov.get("style") == "direct"
        providers_status[pid] = {
            "label": prov["label"],
            "portal": prov["portal"],
            "style": prov.get("style", ""),
            "configured": is_direct or bool(sect.get("cookies")),
            "cookie_count": 0 if is_direct else len(sect.get("cookies", {})),
            "saved_at": sect.get("saved_at", ""),
        }
    active = cfg.get("provider", "pku")
    return {
        "provider": active,
        "active_label": PROVIDERS[active]["label"],
        "providers": providers_status,
        "configured": providers_status[active]["configured"],
        "cookie_count": providers_status[active]["cookie_count"],
        "base_url": PROVIDERS[active]["host"],
    }


@app.post("/api/phase2/auto-login")
async def auto_login_webvpn(data: dict | None = None):
    """Launch Chrome via Playwright for automatic WebVPN login.

    Opens a real Chrome window navigated to the PKU library WebVPN portal
    (wpn.pku.edu.cn). User completes iaaa login manually in that window.
    Cookies are captured automatically and saved to config.
    """
    from .phase2_webvpn import login_via_browser

    provider = (data or {}).get("provider", "")
    result = await login_via_browser(provider=provider) if provider else await login_via_browser()
    if result["success"]:
        return {
            "status": "ok",
            "provider": result.get("provider", ""),
            "cookie_count": result["cookies_count"],
            "message": result["message"],
        }
    else:
        return {
            "status": "error",
            "provider": result.get("provider", ""),
            "cookie_count": 0,
            "message": result["message"],
        }


@app.post("/api/phase2/test")
async def test_phase2_session():
    """Probe the active provider's WebVPN session freshness."""
    from .phase2_webvpn import check_session, get_active_provider

    provider = get_active_provider()
    state = check_session(provider)
    return {"provider": provider, "session": state}


@app.post("/api/phase2/start")
async def start_phase2():
    """Start Phase 2 (library WebVPN) on DOIs that Phase 1 couldn't download."""
    global phase2_running

    if phase2_running:
        return {"error": "Phase 2 is already running"}

    if pipeline_running:
        return {"error": "Phase 1 is still running. Wait for it to finish first."}

    from .phase2_webvpn import get_active_provider, PROVIDERS, get_provider_cookies

    provider = get_active_provider()
    prov = PROVIDERS[provider]
    # direct mode (campus VPN client) authenticates by IP — no cookies required
    if prov.get("style") != "direct" and not get_provider_cookies(provider):
        return {"error": f"No {prov['label']} cookies configured. Run auto-login first."}

    output_dir = str(_get_output_dir())

    from .pipeline import _load_phase1_results
    out_root = Path(output_dir)
    saved = _load_phase1_results(out_root)
    if not saved:
        return {"error": "No Phase 1 results found. Run Phase 1 (Sci-Hub) first."}

    failed = [r for r in saved.get("p1_results", []) if r.get("status") != "downloaded"]
    if not failed:
        return {"error": "All papers were downloaded in Phase 1. No Phase 2 needed."}

    phase2_running = True
    logger.info(f"Starting Phase 2 ({PROVIDERS[provider]['label']}) for {len(failed)} failed DOIs")

    def _run_phase2():
        global phase2_running
        try:
            result = run_phase2_pipeline(
                output_dir=output_dir,
                progress_callback=progress_callback,
            )
            logger.info(f"Phase 2 (WebVPN) finished: {result}")
        except Exception as e:
            logger.error(f"Phase 2 failed: {e}")
            progress_callback("phase2_error", {"error": str(e)})
        finally:
            phase2_running = False

    t = threading.Thread(target=_run_phase2, daemon=True)
    t.start()

    return {"status": "started", "phase": 2, "failed_count": len(failed), "provider": provider}


@app.get("/api/phase2/status")
async def phase2_status():
    """Check Phase 2 status and WebVPN config."""
    from .phase2_webvpn import load_webvpn_config, PROVIDERS

    cfg = load_webvpn_config()
    provider = cfg.get("provider", "pku")
    sect = cfg.get(provider, {})
    # direct mode (campus VPN client) needs no cookies — always ready
    is_direct = PROVIDERS[provider].get("style") == "direct"
    return {
        "running": phase2_running,
        "ready": True,
        "provider": provider,
        "provider_label": PROVIDERS[provider]["label"],
        "style": PROVIDERS[provider].get("style", ""),
        "configured": is_direct or bool(sect.get("cookies")),
    }


# ── Phase 3: MinerU OCR ──────────────────────────────

@app.post("/api/phase3/start")
async def start_phase3():
    """Start Phase 3: MinerU OCR on all downloaded PDFs."""
    global phase3_running

    if phase3_running:
        return {"error": "Phase 3 (MinerU OCR) is already running"}

    if pipeline_running:
        return {"error": "Pipeline (Phase 1+2) is still running. Wait for it to finish first."}

    output_dir = str(_get_output_dir())
    out_root = Path(output_dir)

    # Check that Phase 1+2 has been run (look for results CSV)
    csv_candidates = sorted(
        out_root.glob("results_final_*.csv"),
        key=os.path.getmtime,
        reverse=True,
    )
    if not csv_candidates:
        return {"error": "No results CSV found. Run Phase 1+2 download first."}

    # Check that papers directory has PDFs
    papers_dir = out_root / "papers"
    pdf_files = list(papers_dir.glob("*.pdf")) if papers_dir.exists() else []
    if not pdf_files:
        return {"error": "No PDF files found in papers/ directory. Download some first."}

    # Load MinerU config
    mineru_exe = getattr(_cfg, "MINERU_EXECUTABLE", r"E:\MinerU\.venv\Scripts\mineru.exe") if _cfg else r"E:\MinerU\.venv\Scripts\mineru.exe"
    mineru_backend = getattr(_cfg, "MINERU_BACKEND", "pipeline") if _cfg else "pipeline"
    ocr_subdir = getattr(_cfg, "MINERU_OUTPUT_SUBDIR", "ocr_output") if _cfg else "ocr_output"

    # Check for user-configured MinerU output directory
    ws_cfg = _load_workspace_config()
    mineru_output_dir = ws_cfg.get("mineru_output_dir", "").strip()

    if not Path(mineru_exe).exists():
        return {"error": f"MinerU not found at {mineru_exe}. Check config.py MINERU_EXECUTABLE."}

    phase3_running = True
    logger.info(f"Starting Phase 3 (MinerU OCR) with {len(pdf_files)} PDFs, output → {mineru_output_dir or ocr_subdir}")

    def _run_phase3():
        global phase3_running
        try:
            result = run_phase3(
                results_csv_path="",
                output_dir=output_dir,
                mineru_exe=mineru_exe,
                mineru_backend=mineru_backend,
                ocr_subdir=ocr_subdir,
                ocr_output_dir=mineru_output_dir,
                progress_callback=progress_callback,
            )
            logger.info(f"Phase 3 finished: {result}")
            # Persist phase3 complete flag for reconnection
            state = _load_state()
            state["phase3_complete"] = True
            _save_state(state)
        except Exception as e:
            logger.error(f"Phase 3 failed: {e}")
            progress_callback("phase3_error", {"error": str(e)})
        finally:
            phase3_running = False

    t = threading.Thread(target=_run_phase3, daemon=True)
    t.start()

    return {
        "status": "started",
        "pdfs_to_process": len(pdf_files),
        "results_csv": str(csv_candidates[0]),
    }


@app.get("/api/phase3/status")
async def phase3_status():
    """Check if Phase 3 is running and get latest Excel path."""
    mineru_dir = _get_mineru_output_dir()
    excel_candidates = sorted(
        mineru_dir.glob("results_final_with_ocr_*.xlsx"),
        key=os.path.getmtime,
        reverse=True,
    ) if mineru_dir.exists() else []
    return {
        "running": phase3_running,
        "latest_excel": str(excel_candidates[0]) if excel_candidates else None,
        "latest_excel_name": excel_candidates[0].name if excel_candidates else None,
    }


@app.get("/api/phase3/download")
async def download_excel():
    """Download the latest Phase 3 Excel file."""
    mineru_dir = _get_mineru_output_dir()
    excel_candidates = sorted(
        mineru_dir.glob("results_final_with_ocr_*.xlsx"),
        key=os.path.getmtime,
        reverse=True,
    ) if mineru_dir.exists() else []
    if not excel_candidates:
        return {"error": "No Excel file found. Run Phase 3 first."}
    filepath = excel_candidates[0]
    return FileResponse(str(filepath), filename=filepath.name)


@app.websocket("/ws/progress")
async def websocket_progress(ws: WebSocket):
    """WebSocket endpoint for real-time progress updates."""
    await ws.accept()
    active_connections.append(ws)

    # Send current state on connect (self-heal stale running status first)
    state = _load_state_healed()
    await ws.send_json({"type": "connected", "data": {}, "state": state})

    try:
        while True:
            # Keep connection alive, receive pings
            data = await ws.receive_text()
            if data == "ping":
                await ws.send_json({"type": "pong"})
    except WebSocketDisconnect:
        pass
    finally:
        if ws in active_connections:
            active_connections.remove(ws)


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """Download a result file."""
    out_dir = _get_output_dir()
    filepath = out_dir / filename
    # Also fall back to papers subdirectory
    if not filepath.exists():
        filepath = out_dir / "papers" / filename
    if not filepath.exists():
        return {"error": "File not found"}
    return FileResponse(str(filepath), filename=filename)


# ── Phase 4: LLM Screening ────────────────────────────

@app.post("/api/phase4/start")
async def start_phase4(data: dict):
    """
    Start Phase 4: LLM-based full-text screening.

    Expects JSON body:
      - inclusion_criteria: str
      - exclusion_criteria: str
      - api_key: str (optional, falls back to config.py DEEPSEEK_API_KEY)
    """
    global phase4_running

    if phase4_running:
        return {"error": "Phase 4 (LLM Screening) is already running"}

    if pipeline_running:
        return {"error": "Pipeline (Phase 1+2) is still running. Wait for it to finish."}

    if phase3_running:
        return {"error": "Phase 3 (MinerU OCR) is still running. Wait for it to finish."}

    # ── Read parameters ──
    inclusion_criteria = data.get("inclusion_criteria", "").strip()
    exclusion_criteria = data.get("exclusion_criteria", "").strip()
    web_api_key = data.get("api_key", "").strip()
    web_model = data.get("model", "").strip()
    web_api_base = data.get("api_base", "").strip()
    web_max_chars = data.get("max_content_chars")
    web_temperature = data.get("temperature")

    if not inclusion_criteria and not exclusion_criteria:
        # Default criteria are built into phase4_screening.py — proceed with defaults
        logger.info("Phase 4: No custom criteria provided — using built-in 4-step decision tree defaults.")

    # ── Resolve parameters: web UI > workspace_config.json > config.py > defaults ──
    ws_cfg = _load_workspace_config()
    api_key = web_api_key or ws_cfg.get("deepseek_api_key", "") or (getattr(_cfg, "DEEPSEEK_API_KEY", "") if _cfg else "")
    if not api_key:
        return {
            "error": (
                "DeepSeek API key not configured. "
                "Set it in the Web UI or in config.py (DEEPSEEK_API_KEY)."
            )
        }

    # ── Resolve remaining config values ──
    api_base = web_api_base or ws_cfg.get("deepseek_api_base", "") or (getattr(_cfg, "DEEPSEEK_BASE_URL", "https://api.deepseek.com") if _cfg else "https://api.deepseek.com")
    model = web_model or ws_cfg.get("deepseek_model", "") or (getattr(_cfg, "DEEPSEEK_MODEL", "deepseek-v4-pro") if _cfg else "deepseek-v4-pro")
    max_chars = int(web_max_chars) if web_max_chars is not None else (ws_cfg.get("deepseek_max_content_chars") or ws_cfg.get("screening_max_chars") or (getattr(_cfg, "SCREENING_MAX_CHARS", 90000) if _cfg else 90000))
    temperature = float(web_temperature) if web_temperature is not None and web_temperature != "" else (ws_cfg.get("deepseek_temperature") or ws_cfg.get("screening_temperature") or (getattr(_cfg, "SCREENING_TEMPERATURE", 0.1) if _cfg else 0.1))
    api_delay = getattr(_cfg, "SCREENING_API_DELAY", 2.0) if _cfg else 2.0

    output_dir = str(_get_output_dir())
    mineru_output_dir = str(_get_mineru_output_dir())

    # ── Save criteria to workspace config for persistence ──
    ws_cfg["screening_inclusion"] = inclusion_criteria
    ws_cfg["screening_exclusion"] = exclusion_criteria
    if web_api_key:
        ws_cfg["deepseek_api_key"] = web_api_key
    if web_model:
        ws_cfg["deepseek_model"] = web_model
    if web_api_base:
        ws_cfg["deepseek_api_base"] = web_api_base
    ws_cfg["screening_max_chars"] = max_chars
    ws_cfg["screening_temperature"] = temperature
    _save_workspace_config(ws_cfg)

    phase4_running = True
    logger.info(
        f"Starting Phase 4 (LLM Screening) | model={model} | "
        f"inclusion_len={len(inclusion_criteria)} | exclusion_len={len(exclusion_criteria)}"
    )

    def _run_phase4():
        global phase4_running
        try:
            result = run_phase4(
                inclusion_criteria=inclusion_criteria,
                exclusion_criteria=exclusion_criteria,
                excel_path="",
                output_dir=output_dir,
                mineru_output_dir=mineru_output_dir,
                api_key=api_key,
                api_base=api_base,
                model=model,
                max_content_chars=max_chars,
                temperature=temperature,
                api_delay=api_delay,
                progress_callback=progress_callback,
            )
            logger.info(f"Phase 4 finished: {result}")
            # Persist phase4 complete flag for reconnection
            state = _load_state()
            state["phase4_complete"] = True
            _save_state(state)
        except Exception as e:
            logger.error(f"Phase 4 failed: {e}")
            progress_callback("phase4_error", {"error": str(e)})
        finally:
            phase4_running = False

    t = threading.Thread(target=_run_phase4, daemon=True)
    t.start()

    # Count papers with OCR-ready MD files for the response
    # (same logic as run_phase4: filename-stem match + DOI-based fallback)
    import openpyxl as _xl
    from .phase4_screening import _build_md_index, _clean_doi
    mineru_dir = _get_mineru_output_dir()
    md_index = _build_md_index(mineru_dir)
    ocr_ready = 0
    try:
        excel_candidates = sorted(
            list(Path(mineru_dir).glob("results_final_with_ocr_*.xlsx")),
            key=os.path.getmtime, reverse=True,
        )
        if excel_candidates:
            wb = _xl.load_workbook(excel_candidates[0])
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            col_map = {h.lower().replace(" ", "_"): i + 1 for i, h in enumerate(headers) if h is not None}
            doi_col = col_map.get("doi")
            fn_col = col_map.get("filename")
            if doi_col and fn_col:
                for row_idx in range(2, ws.max_row + 1):
                    doi = str(ws.cell(row=row_idx, column=doi_col).value or "").strip()
                    fname = str(ws.cell(row=row_idx, column=fn_col).value or "").strip()
                    if not doi:
                        continue
                    # Try filename stem first, then DOI-based
                    found = False
                    if fname:
                        stem = Path(fname).stem
                        found = stem in md_index
                    if not found:
                        doi_clean = _clean_doi(doi)
                        doi_slug = doi_clean.replace("/", "_")
                        for c in (f"manual_{doi_slug}", doi_slug):
                            if c in md_index:
                                found = True
                                break
                        if not found:
                            # Step B: case-insensitive prefix match
                            search = f"manual_{doi_slug}".lower()[:60]
                            for k in md_index:
                                if k.lower().startswith(search):
                                    found = True
                                    break
                    if found:
                        ocr_ready += 1
    except Exception:
        pass

    return {
        "status": "started",
        "ocr_ready_papers": ocr_ready,
        "model": model,
    }


@app.get("/api/phase4/status")
async def phase4_status():
    """Check if Phase 4 is running and get latest screened Excel path."""
    out_root = _get_output_dir()
    excel_candidates = sorted(
        out_root.glob("results_final_screened_*.xlsx"),
        key=os.path.getmtime,
        reverse=True,
    ) if out_root.exists() else []
    return {
        "running": phase4_running,
        "latest_excel": str(excel_candidates[0]) if excel_candidates else None,
        "latest_excel_name": excel_candidates[0].name if excel_candidates else None,
    }


@app.get("/api/phase4/download")
async def download_screened_excel():
    """Download the latest Phase 4 screened Excel file."""
    out_root = _get_output_dir()
    excel_candidates = sorted(
        out_root.glob("results_final_screened_*.xlsx"),
        key=os.path.getmtime,
        reverse=True,
    ) if out_root.exists() else []
    if not excel_candidates:
        return {"error": "No screened Excel file found. Run Phase 4 first."}
    filepath = excel_candidates[0]
    return FileResponse(str(filepath), filename=filepath.name)


@app.get("/api/files")
async def list_files():
    """List all output files."""
    out_dir = _get_output_dir()
    outputs = []
    # CSV + Excel files in output root
    if out_dir.exists():
        for f in sorted(out_dir.glob("*"), key=os.path.getmtime, reverse=True):
            if f.is_file():
                outputs.append({"name": f.name, "size": f.stat().st_size, "modified": f.stat().st_mtime})
    # PDF files in papers/
    papers_dir = out_dir / "papers"
    if papers_dir.exists():
        for f in sorted(papers_dir.glob("*.pdf"), key=os.path.getmtime, reverse=True):
            outputs.append({"name": f"papers/{f.name}", "size": f.stat().st_size, "modified": f.stat().st_mtime})
    return {"outputs": outputs, "output_dir": str(out_dir)}


# ── Results Management: review + manual PDF supplement ──

@app.get("/api/results")
async def get_results():
    """Return the latest pipeline results as JSON (all papers with full status).

    Checks results_final_*.csv first. If missing or stale, falls back to
    phase1_results.json (which is written when Phase 1 has failures and
    pauses for Phase 2).
    """
    out_root = _get_output_dir()

    # ── Try phase1_results.json first if it's newer/larger than CSV ──
    p1_json = out_root / "phase1_results.json"
    csv_path = _find_latest_results_csv(out_root)

    use_json = False
    if p1_json.exists():
        if not csv_path:
            use_json = True
        else:
            # Compare modification times; also check row count
            json_mtime = os.path.getmtime(p1_json)
            csv_mtime = os.path.getmtime(csv_path)
            if json_mtime > csv_mtime:
                use_json = True
            else:
                # CSV exists and is newer — but might be stale test data
                csv_papers = _read_results_csv(csv_path)
                if len(csv_papers) < 20:
                    # Very small CSV, probably from a test run — prefer JSON
                    use_json = True

    papers = []
    source_name = ""
    source_path = ""

    if use_json:
        try:
            with open(p1_json, "r", encoding="utf-8") as f:
                p1_data = json.load(f)
            raw = p1_data.get("p1_results", [])
            # Normalize field names for frontend compatibility
            for r in raw:
                p = {
                    "doi": r.get("doi", ""),
                    "title": r.get("title", ""),
                    "filename": r.get("filename", ""),
                    "status": r.get("status", ""),
                    "phase": "1",
                    "filepath": r.get("filepath", ""),
                    "message": r.get("message", ""),
                    "route": r.get("route", ""),
                }
                # Enrich with CSV metadata (first_author, year) if available
                papers.append(p)
            source_name = "phase1_results.json"
            source_path = str(p1_json)
        except Exception as e:
            logger.error(f"Failed to read phase1_results.json: {e}")
            if csv_path:
                papers = _read_results_csv(csv_path)
                source_name = csv_path.name
                source_path = str(csv_path)
    elif csv_path:
        papers = _read_results_csv(csv_path)
        source_name = csv_path.name
        source_path = str(csv_path)

    if not papers:
        return {"papers": [], "summary": {
            "total": 0, "downloaded": 0, "missing": 0,
            "status_counts": {}, "csv_name": "", "csv_path": "",
        }}

    # ── Enrich with metadata from original CSV (first_author, year, etc.) ──
    orig_csv = out_root.parent / "phase1_prior_reviews_updated.csv"
    if not orig_csv.exists():
        # Try the CSV path stored in phase1_results.json
        if use_json and p1_json.exists():
            try:
                with open(p1_json, "r", encoding="utf-8") as f:
                    p1_data = json.load(f)
                orig_csv_str = p1_data.get("csv_path", "")
                if orig_csv_str:
                    orig_csv = Path(orig_csv_str)
            except Exception:
                pass

    orig_meta = {}
    if orig_csv.exists():
        try:
            orig_papers = _read_results_csv(orig_csv)
            for p in orig_papers:
                doi = p.get("doi", "").strip().lower()
                if doi:
                    orig_meta[doi] = p
        except Exception:
            pass

    for p in papers:
        doi = p.get("doi", "").strip().lower()
        if doi in orig_meta:
            meta = orig_meta[doi]
            if not p.get("first_author"):
                p["first_author"] = meta.get("first_author", "")
            if not p.get("year"):
                p["year"] = meta.get("year", "")
            if not p.get("title"):
                p["title"] = meta.get("title", "")

    # ── Build summaries ──
    status_counts: dict[str, int] = {}
    for p in papers:
        s = p.get("status", "unknown").strip()
        status_counts[s] = status_counts.get(s, 0) + 1

    has_pdf_statuses = {"downloaded", "already_downloaded", "manual_added"}
    has_pdf = sum(1 for p in papers if p.get("status", "").strip() in has_pdf_statuses)
    missing = len(papers) - has_pdf

    summary = {
        "total": len(papers),
        "downloaded": has_pdf,
        "missing": missing,
        "coverage_pct": round(has_pdf / len(papers) * 100, 1) if papers else 0,
        "status_counts": status_counts,
        "csv_name": source_name,
        "csv_path": source_path,
    }

    return {"papers": papers, "summary": summary}


@app.post("/api/results/upload-pdf")
async def upload_pdf_for_doi(
    doi: str = Form(...),
    file: UploadFile = File(...),
):
    """Upload a PDF file for a specific DOI (for papers not auto-downloaded).

    The PDF is saved to the papers/ directory and the results CSV is updated
    with status 'manual_added'.
    """
    out_root = _get_output_dir()
    papers_dir = out_root / "papers"
    papers_dir.mkdir(parents=True, exist_ok=True)

    # Validate
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        return {"error": "Only PDF files are accepted."}

    doi_clean = doi.strip().lower()
    if not doi_clean:
        return {"error": "DOI is required."}

    # Check against results CSV
    csv_path = _find_latest_results_csv(out_root)
    if not csv_path:
        return {"error": "No results CSV found. Run Phase 1+2 first."}

    papers = _read_results_csv(csv_path)
    matched = None
    for p in papers:
        if p.get("doi", "").strip().lower() == doi_clean:
            matched = p
            break

    # Generate filename from paper metadata (Author_Year_Title.pdf)
    target_name = _build_paper_filename(matched or {"doi": doi_clean}, papers_dir)
    target_path = papers_dir / target_name

    # Save uploaded file
    content = await file.read()
    with open(target_path, "wb") as f:
        f.write(content)

    # Update results CSV
    _update_paper_status(out_root, doi_clean, "manual_added", str(target_path))

    # Notify frontend via WebSocket
    if matched:
        matched["status"] = "manual_added"
        matched["filepath"] = str(target_path)
        matched["filename"] = target_name
        progress_callback("results_updated", {
            "doi": doi_clean,
            "action": "pdf_uploaded",
            "filename": target_name,
            "total_downloaded": sum(
                1 for p in papers if p.get("status", "") in
                {"downloaded", "already_downloaded", "manual_added"}
            ),
        })

    return {
        "status": "uploaded",
        "doi": doi_clean,
        "filename": target_name,
        "filepath": str(target_path),
        "size": len(content),
    }


@app.post("/api/results/import-local-pdf")
async def import_local_pdf(data: dict):
    """Import a PDF from a local file path for a specific DOI.

    Expects JSON: {"doi": "...", "path": "/absolute/path/to/file.pdf"}
    The file is copied (not moved) to papers/ and the results CSV is updated.
    """
    doi = data.get("doi", "").strip()
    local_path = data.get("path", "").strip()

    if not doi or not local_path:
        return {"error": "Both 'doi' and 'path' are required."}

    source = Path(local_path)
    if not source.exists():
        return {"error": f"File not found: {local_path}"}
    if source.suffix.lower() != ".pdf":
        return {"error": "Only PDF files are accepted."}

    out_root = _get_output_dir()
    papers_dir = out_root / "papers"
    papers_dir.mkdir(parents=True, exist_ok=True)

    # Check against results CSV
    csv_path = _find_latest_results_csv(out_root)
    if not csv_path:
        return {"error": "No results CSV found. Run Phase 1+2 first."}

    doi_clean = doi.lower()
    papers = _read_results_csv(csv_path)
    matched = None
    for p in papers:
        if p.get("doi", "").strip().lower() == doi_clean:
            matched = p
            break

    # Generate filename from paper metadata (Author_Year_Title.pdf)
    target_name = _build_paper_filename(matched or {"doi": doi_clean}, papers_dir)
    target_path = papers_dir / target_name

    # Copy
    shutil.copy2(source, target_path)

    # Update
    _update_paper_status(out_root, doi_clean, "manual_added", str(target_path))

    progress_callback("results_updated", {
        "doi": doi_clean,
        "action": "pdf_imported",
        "filename": target_name,
        "total_downloaded": sum(
            1 for p in papers if p.get("status", "") in
            {"downloaded", "already_downloaded", "manual_added"}
        ),
    })

    return {
        "status": "imported",
        "doi": doi_clean,
        "source": str(source),
        "filename": target_name,
        "filepath": str(target_path),
    }


@app.get("/api/results/export-excel")
async def export_results_excel():
    """Download results as a formatted Excel (.xlsx) with colored status rows."""
    out_root = _get_output_dir()
    csv_path = _find_latest_results_csv(out_root)
    if not csv_path:
        return {"error": "No results found. Run Phase 1+2 first."}

    papers = _read_results_csv(csv_path)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Download Results"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    manual_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    headers = ["DOI", "Title", "First Author", "Year", "Journal",
               "Filename", "Status", "Phase", "Filepath", "Message"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, p in enumerate(papers, 2):
        values = [
            p.get("doi", ""),
            p.get("title", ""),
            p.get("first_author", ""),
            p.get("year", ""),
            p.get("journal", ""),
            p.get("filename", ""),
            p.get("status", ""),
            p.get("phase", ""),
            p.get("filepath", ""),
            p.get("message", ""),
        ]
        status = p.get("status", "").strip()
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")

        # Color row by status
        if status in ("downloaded", "already_downloaded"):
            row_fill = success_fill
        elif status == "manual_added":
            row_fill = manual_fill
        else:
            row_fill = missing_fill

        for c in range(1, 11):
            ws.cell(row=row_idx, column=c).fill = row_fill

    # Column widths
    widths = [32, 55, 16, 8, 22, 42, 16, 7, 55, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A1:J{len(papers) + 1}"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    has_pdf_set = {"downloaded", "already_downloaded", "manual_added"}
    total = len(papers)
    downloaded = sum(1 for p in papers if p.get("status", "").strip() in has_pdf_set)

    summary_data = [
        ("Total Papers", total),
        ("Full Text Available", downloaded),
        ("Full Text Missing", total - downloaded),
        ("Coverage Rate", f"{downloaded / total * 100:.1f}%" if total > 0 else "0%"),
        ("", ""),
        ("--- Status Breakdown ---", ""),
    ]
    status_counts: dict[str, int] = {}
    for p in papers:
        s = p.get("status", "").strip()
        if s:
            status_counts[s] = status_counts.get(s, 0) + 1
    for st, cnt in sorted(status_counts.items()):
        summary_data.append((st, cnt))

    for i, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 14

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = out_root / f"results_summary_{timestamp}.xlsx"
    wb.save(xlsx_path)
    logger.info(f"Exported results Excel: {xlsx_path}")

    return FileResponse(str(xlsx_path), filename=xlsx_path.name)


@app.get("/api/phase2/export-missing")
async def export_missing_fulltext():
    """Export papers WITHOUT full text as a separate Excel file.

    Reads the latest results CSV, filters papers where status is NOT in
    the "has PDF" set (downloaded/already_downloaded/manual_added), and
    generates a standalone Excel with two sheets:
      - Missing Papers (detailed)
      - Summary (counts by status)

    This endpoint does NOT modify the Phase 2 CSV — it only reads and exports.
    """
    out_root = _get_output_dir()
    csv_path = _find_latest_results_csv(out_root)
    if not csv_path:
        return {"error": "No results found. Run Phase 1+2 first."}

    papers = _read_results_csv(csv_path)
    if not papers:
        return {"error": "Results CSV is empty."}

    has_pdf_set = {"downloaded", "already_downloaded", "manual_added"}

    # Filter: papers WITHOUT full text
    missing_papers = [
        p for p in papers
        if p.get("status", "").strip() not in has_pdf_set
    ]

    if not missing_papers:
        return {"error": "All papers have full text — nothing to export."}

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missing Papers"

    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    no_pdf_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    idx_font = Font(bold=True, size=10)

    # ── Headers ──
    headers = ["#", "DOI / ID", "Title", "First Author", "Year",
               "Journal / Source", "Status", "Message"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ── Data rows ──
    for i, p in enumerate(missing_papers):
        row = i + 2
        values = [
            i + 1,
            p.get("doi", ""),
            p.get("title", ""),
            p.get("first_author", ""),
            p.get("year", ""),
            p.get("journal", ""),
            p.get("status", ""),
            p.get("message", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3, 6, 8)))
            if col_idx == 1:
                cell.font = idx_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = no_pdf_fill

    # ── Column widths ──
    col_widths = {"A": 5, "B": 32, "C": 55, "D": 16, "E": 8, "F": 26, "G": 18, "H": 40}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze header + auto-filter ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(missing_papers) + 1}"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    status_counts: dict[str, int] = {}
    for p in missing_papers:
        s = p.get("status", "").strip()
        if s:
            status_counts[s] = status_counts.get(s, 0) + 1

    ws2.cell(row=1, column=1, value="Missing Full-Text Papers — Summary").font = Font(bold=True, size=13)

    total = len(papers)
    missing_total = len(missing_papers)
    summary_data = [
        ("", ""),
        ("Total papers in CSV", total),
        ("Papers with full text", total - missing_total),
        ("Papers WITHOUT full text", missing_total),
        ("Coverage rate", f"{(total - missing_total) / total * 100:.1f}%" if total > 0 else "0%"),
        ("", ""),
        ("--- Status Breakdown ---", ""),
    ]
    for i, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)

    row = len(summary_data) + 4
    for st, cnt in sorted(status_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=1, value=st)
        ws2.cell(row=row, column=2, value=cnt)
        row += 1

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14

    # Add note about source
    note_row = row + 2
    ws2.cell(row=note_row, column=1, value="Source CSV:").font = Font(bold=True, color="888888")
    ws2.cell(row=note_row, column=2, value=csv_path.name)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = out_root / f"missing_fulltext_{timestamp}.xlsx"
    wb.save(xlsx_path)
    logger.info(f"Exported missing full-text list: {xlsx_path} ({missing_total} papers)")

    return FileResponse(str(xlsx_path), filename=xlsx_path.name)


# Serve frontend static files
if FRONTEND_DIR.exists():
    @app.get("/")
    async def serve_frontend():
        dashboard = FRONTEND_DIR / "dashboard.html"
        if dashboard.exists():
            return FileResponse(str(dashboard))
        return {"message": "DoiHarvest API", "docs": "/docs"}
else:
    @app.get("/")
    async def root():
        return {"message": "DoiHarvest API", "docs": "/docs"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8765)
