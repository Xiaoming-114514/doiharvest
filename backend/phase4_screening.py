"""
Phase 4: LLM-based literature screening for scoping review.

Reads the Phase 3 Excel, sends each paper's extracted Markdown to DeepSeek API
with a built-in 4-step decision tree, and produces a final Excel with:

  - screening_decision  (Include / Exclude / Uncertain / Skip / Error)
  - screening_reason    (per-step evidence + overall reasoning in Chinese)

The 4-step decision tree is embedded in SYSTEM_PROMPT + SCREENING_USER_PROMPT:
  STEP 1 — Document type gate (must be original research)
  STEP 2 — Population framework (SSD / somatoform / MUS / bodily distress)
  STEP 3 — Assessment tool check (named clinical tool required)
  STEP 4 — Study design check (acceptable designs only)

Key design decisions:
  - Temperature 0.1 for reproducible Include/Exclude decisions
  - System message + user message structure (better instruction-following)
  - Semantic (concept-based) matching, NOT keyword matching
  - Bilingual signal words (EN + ZH) for cross-language papers
  - Uncertain decision for low-confidence papers (flagged for human review)
  - Default criteria built-in; users can optionally add custom criteria via Web UI
"""

import json
import logging
import os
import re
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
StatusCallback = Callable[[str, dict], None] | None

# ── LLM client cache ──
_client = None
_client_config = {}


def _get_client(api_key: str, base_url: str = "https://api.deepseek.com"):
    """Get or create cached OpenAI-compatible client."""
    global _client, _client_config
    new_cfg = {"api_key": api_key, "base_url": base_url}
    if _client is not None and _client_config == new_cfg:
        return _client
    from openai import OpenAI
    # timeout=120: 防止 DeepSeek 偶发挂起时单个调用卡 10 分钟（SDK 默认）+ 重试，
    # 导致整个 Phase 4 长时间停滞。120 秒对全文筛选足够。
    _client = OpenAI(api_key=api_key, base_url=base_url, timeout=120)
    _client_config = new_cfg
    return _client


# ── Prompt Templates ─────────────────────────────────

SYSTEM_PROMPT = """You are an expert systematic review screener for a staged scoping review on clinical assessment tools for somatic symptom disorder (SSD) across diagnostic transitions.

Your task: read the full text of a research paper (extracted from PDF via OCR) and determine whether it should be INCLUDED or EXCLUDED using a 4-step decision tree.

CRITICAL PRINCIPLES:
1. CONCEPTUAL understanding, NOT keyword matching. For example: "patients with multiple physical complaints that remain medically unexplained after extensive workup" is semantically equivalent to "medically unexplained symptoms (MUS)" — even if the exact acronym MUS never appears. Judge the meaning, not the string.
2. The paper may be in English or Chinese. Process either language naturally. The signal terms listed in the criteria are bilingual guides, not an exhaustive checklist.
3. When your confidence is below 80% on any step, output "Uncertain" for that step — do NOT guess. It is better to flag for human review than to make a wrong call.
4. Cite specific evidence from the paper (quoted terms, section references, study descriptions) in each step's result.
5. This is a SCOPING REVIEW (not an intervention meta-analysis). The inclusion criteria focus on POPULATION FRAMEWORK, PRESENCE OF ASSESSMENT TOOLS, and ACCEPTABLE STUDY DESIGNS. There is NO requirement for specific interventions, comparators, or clinical outcomes beyond the assessment tools themselves."""


SCREENING_USER_PROMPT = """## 纳入/排除标准 — 4 步决策树

请按以下 4 个步骤逐条评估这篇论文。理解概念含义，不要做关键词匹配。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STEP 1 — 文献类型初筛 (Document Type Gate)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
判断本文是否为原始研究（original research）。

✓ 可纳入类型：
  cross-sectional / 横断面, case-control / 病例对照, cohort / 队列,
  scale development or validation / 量表开发或验证,
  diagnostic accuracy / 诊断准确性,
  RCT / 随机对照试验, intervention study / 干预研究,
  health services research / 卫生服务研究

✗ 排除：
  - Review / Meta-analysis / 综述（已在 Phase 1 单独处理）
  - Case report / 病例报告（N=1）
  - Commentary / Editorial / Letter / Conference abstract / 评论 / 社论 / 会议摘要

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STEP 2 — 人群框架检查 (Population Framework)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
判断研究对象是否属于躯体症状障碍谱系。
核心问题：研究人群是否以躯体症状为临床焦点，且被置于 SSD / somatoform / MUS / bodily distress 的诊断或概念框架下？

EN semantic signals（任一情境即可，不要求精确匹配术语）:
  "somatic symptom disorder" / SSD / "DSM-5 somatic symptom"
  "bodily distress disorder" / BDD / "bodily distress syndrome" / BDS
  "somatoform disorder" / "somatization disorder" / "Briquet syndrome"
  "somatization" / "somatisation"
  "medically unexplained symptoms" / MUS / "medically unexplained physical symptoms" / MUPS
  描述为 "persistent physical symptoms without organic explanation"
  描述为 "multiple somatic complaints" / "functional somatic syndromes"

ZH semantic signals（任一情境即可）:
  "躯体症状障碍" / "躯体形式障碍" / "躯体化障碍"
  "躯体化" / "医学无法解释的症状"
  "躯体痛苦障碍" / "躯体忧虑障碍" / "Briquet 综合征"
  描述为 "反复就诊、检查阴性" / "无明显器质性基础" / "功能性躯体不适"

条件纳入：IBS / fibromyalgia / CFS / 功能性消化不良 / 慢性盆腔痛 / 慢性疼痛 等功能性躯体综合征，仅当原文明确将其置于 somatization / MUS / SSD / bodily distress 框架下时通过。

✗ 排除：
  - 仅研究抑郁/焦虑/精神分裂症中的躯体症状，未置于 SSD/somatoform/MUS 框架
  - 纯专科疾病研究（IBS/CFS/fibromyalgia等），全文未提及 somatization/MUS/SSD

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STEP 3 — 评估工具检查 (Assessment Tool Check)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
论文是否报告了至少一个命名的临床评估工具或临床测量指标？

工具包括以下任一类型：
  A) 标准化量表/问卷：PHQ-15, SSS-8, SSD-12, SOMS, Whiteley Index, SHAI, HAI,
     BDS checklist, WHODAS, SF-36, EQ-5D, SCL-90, BSI, GAD-7, PHQ-9,
     或任何有名称的心理测量工具
  B) 结构化诊断访谈：SCID, MINI, CIDI, 或基于 DSM/ICD 的结构化访谈
  C) 研究者自编/改编工具（需有条目内容或评分方式描述）
  D) 医疗使用指标作为研究变量（就诊次数/转诊/费用等）

✗ 排除：
  - 仅报告 fMRI/EEG/血液标志物/基因/炎症因子/cortisol 等纯生物学指标，无临床评估工具
  - 工具仅作为 Table 1 基线描述变量，Methods 中未说明其作为研究变量的测量目的

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STEP 4 — 研究设计检查 (Study Design Check)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
确认研究设计属于可纳入的类型。

✗ 排除：
  - 动物实验 / 细胞实验
  - 纯机制研究（仅探讨病理生理通路，不涉及临床评估）
  - 纯机器学习/算法模型研究（仅报告模型性能，无临床评估工具的验证或应用）

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
综合判定规则
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• STEP 1–4 全部通过 → Include
• 任一步不通过 → Exclude（注明触发步骤及论文中的证据）
• 任一步信息不足、置信度 < 80% → Uncertain（说明哪个步骤不确定及原因）

{extra_criteria}

## 论文全文（由 MinerU 从 PDF 提取的 Markdown）

---
{paper_content}
---

## 输出格式（严格 JSON，不要包含 ```json 标记或其他文本）

{{
  "decision": "Include",
  "step1_doc_type": "Pass — <证据>",
  "step2_population": "Pass — <引述论文中的框架术语或人群描述>",
  "step3_tool": "Pass — <论文使用的评估工具名称>",
  "step4_design": "Pass — <识别的研究设计>",
  "reason": "综合理由（中文）。该研究满足所有纳入标准：研究对象为...（论文描述），使用了...工具，研究设计为...。"
}}

或：

{{
  "decision": "Exclude",
  "step1_doc_type": "Pass/Fail — <证据>",
  "step2_population": "Pass/Fail — <证据>",
  "step3_tool": "Pass/Fail — <证据>",
  "step4_design": "Pass/Fail — <证据>",
  "reason": "排除原因（中文）。明确触发排除的具体步骤编号及证据。例如：'STEP 2 排除：该研究纳入慢性下背痛患者，全文未提及 somatization/MUS/SSD 框架，不满足本研究 PCC 人群要求。'"
}}

或：

{{
  "decision": "Uncertain",
  "step1_doc_type": "Pass/Fail/Uncertain — <证据>",
  "step2_population": "Pass/Fail/Uncertain — <证据>",
  "step3_tool": "Pass/Fail/Uncertain — <证据>",
  "step4_design": "Pass/Fail/Uncertain — <证据>",
  "reason": "不确定原因（中文）。说明哪个步骤不确定及原因。例如：'STEP 2 不确定：论文纳入 primary care patients with persistent physical symptoms，症状描述符合 MUS 特征但作者未使用 somatization/MUS/SSD 等框架术语来定义人群，需��工复核。'"
}}

现在请评估这篇论文，输出 JSON："""


CUSTOM_SCREENING_USER_PROMPT = """## 文献筛选任务

请根据以下**用户自定义**的纳入/排除标准，判断这篇论文应纳入还是排除。
严格依据这些标准判断，不要套用任何其他默认规则。

## 用户纳入标准
{inclusion_criteria}

## 用户排除标准
{exclusion_criteria}

## 判定规则
- 论文满足纳入标准、且不触发任何排除标准 → Include
- 论文触发任一排除标准、或不满足纳入标准 → Exclude
- 信息不足、无法确定（置信度 < 80%）→ Uncertain

## 论文全文（由 MinerU 从 PDF 提取的 Markdown）

---
{paper_content}
---

## 输出格式（严格 JSON，不要包含 ```json 标记或其他文本）

{{
  "decision": "Include",
  "reason": "判定理由（中文）。明确说明论文满足用户标准的哪些具体条目。"
}}

或：

{{
  "decision": "Exclude",
  "reason": "排除原因（中文）。明确说明触发的具体排除条目或未满足的纳入条目。"
}}

或：

{{
  "decision": "Uncertain",
  "reason": "不确定原因（中文）。说明哪项标准无法判断及原因。"
}}

请评估这篇论文，输出 JSON："""


# ── Default Screening Criteria (for Web UI pre-fill) ──

DEFAULT_INCLUSION_SUMMARY = """本研究是一项关于躯体症状障碍（SSD）临床评估工具的 staged scoping review。

核心纳入条件（PCC 框架）：
• Population：SSD / bodily distress disorder / somatoform / somatization / MUS / MUPS 人群
• Concept：论文中至少使用了一个命名的临床评估工具（量表/问卷/结构化访谈/医疗使用指标）
• Context：横断面/病例对照/队列/量表验证/诊断准确性/RCT/干预/卫生服务研究等原始研究

具体判断标准见 4 步决策树（已内置于筛选 prompt 中）。"""

DEFAULT_EXCLUSION_SUMMARY = """排除：
• 非原始研究（综述/Meta分析/评论/病例报告/会议摘要）
• 人群不匹配（仅抑郁/焦虑/精分中的躯体症状，或纯 IBS/CFS/慢性疼痛专科研究，未置于目标框架下）
• 纯生物学标志物研究（仅 fMRI/EEG/血液/基因），无临床评估工具
• 评估工具仅作为 Table 1 基线协变量，Methods 未说明测量目的
• 动物实验 / 细胞实验 / 纯机制研究 / 纯 ML 模型研究"""


# ── Core Functions ───────────────────────────────────

def _find_latest_phase3_excel(mineru_dir: Path, output_dir: Path) -> Path | None:
    """Find the most recent Phase 3 Excel file."""
    # Search in mineru dir first, then output dir
    for search_dir in (mineru_dir, output_dir):
        if search_dir.exists():
            candidates = sorted(
                list(search_dir.glob("results_final_with_ocr_*.xlsx")),
                key=os.path.getmtime,
                reverse=True,
            )
            if candidates:
                return candidates[0]
    return None


# Per-session cache: stem → Path, built once on first use. Avoids O(N)
# glob rescans for 500+ folders (saves ~30 s per Phase 4 run).
_md_index: dict[str, Path] | None = None


def _build_md_index(mineru_dir: Path) -> dict[str, Path]:
    """Walk the OCR output tree once and build a stem → md_path map.

    Also populates a companion DOI→md_path index for fast fallback lookups.
    """
    index: dict[str, Path] = {}
    for md_path in mineru_dir.glob("*/auto/*.md"):
        index[md_path.stem] = md_path
    # Also catch non-auto layouts (older MinerU versions)
    for md_path in mineru_dir.glob("*/*.md"):
        if md_path.stem not in index:
            index[md_path.stem] = md_path
    return index


def _clean_doi(doi: str) -> str:
    """Strip trailing junk (spaces, parentheses, percent-encoded titles) from a DOI value.

    Some Excel DOI cells contain extra text appended after the DOI proper, e.g.:
        "10.1055/a-1197-6068     %(Die deutschsprachige Fassung..."
    This extracts only the actual DOI prefix.
    """
    if not doi:
        return ""
    doi = doi.strip()
    # Split on the first occurrence of these delimiters (not part of a valid DOI)
    for sep in ("     ", "\t", "\n", "\r", " %(", " % ("):
        idx = doi.find(sep)
        if idx > 0:
            doi = doi[:idx]
            break
    return doi.rstrip(".,;:")


def _find_md_file(mineru_dir: Path, pdf_filename: str) -> Path | None:
    """
    Locate the MinerU output .md file for a given PDF filename.

    Strategy (ordered):
      1. Exact folder match:  {mineru_dir}/{stem}/auto/{stem}.md
      2. No-auto layout:      {mineru_dir}/{stem}/{stem}.md
      3. Pre-built index (covers folder-name mismatches: OCR folder
         was renamed to full title but MD file keeps truncated stem).
      4. Slow glob fallback (only when index is empty — first call).
    """
    global _md_index
    stem = Path(pdf_filename).stem

    # 1. Exact match — folder name == stem
    primary = mineru_dir / stem / "auto" / f"{stem}.md"
    if primary.exists():
        return primary

    # 2. No-auto layout
    alt1 = mineru_dir / stem / f"{stem}.md"
    if alt1.exists():
        return alt1

    # 3. Pre-built index (handles folder-name ≠ stem-name case)
    if _md_index is None:
        _md_index = _build_md_index(mineru_dir)
    if stem in _md_index:
        return _md_index[stem]

    # 4. Slow glob — last resort
    for md_path in mineru_dir.glob(f"**/{stem}.md"):
        return md_path

    # 5. OCR folder renamed but MD file inside kept old name:
    #    {mineru_dir}/{new_stem}/auto/{old_stem}.md
    folder = mineru_dir / stem
    if folder.is_dir():
        auto_mds = list(folder.glob("auto/*.md"))
        if auto_mds:
            return auto_mds[0]
        # Also check top-level MDs
        top_mds = list(folder.glob("*.md"))
        if top_mds:
            return top_mds[0]

    return None


def _parse_llm_response(raw_text: str) -> tuple[str, str]:
    """Parse the LLM JSON response. Returns (decision, reason).

    Handles the decision-tree format with step_results, and falls back
    to the legacy format with only decision + reason.
    """
    # Strip markdown code blocks
    text = raw_text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*\n?", "", text)
        text = re.sub(r"\n?\s*```\s*$", "", text)

    def _normalize_decision(d: str) -> str:
        d = d.strip().lower()
        if "include" in d:
            return "Include"
        elif "exclude" in d:
            return "Exclude"
        elif "uncertain" in d:
            return "Uncertain"
        return "Exclude"

    try:
        result = json.loads(text)
        decision = _normalize_decision(str(result.get("decision", "")))
        reason = str(result.get("reason", "")).strip()

        # If step_results exist, enrich the reason with per-step details
        step_results = result.get("step_results", {})
        if step_results:
            step_summary = []
            for key in ("step1_doc_type", "step2_population", "step3_tool", "step4_design"):
                val = step_results.get(key, "")
                if val:
                    step_summary.append(f"[{key}] {val}")
            if step_summary:
                reason = " | ".join(step_summary) + (" | " + reason if reason else "")

        return decision, reason
    except json.JSONDecodeError:
        # Try to extract JSON from the text — handle nested braces
        # Strategy: find the outermost { ... } that contains "decision"
        try:
            # Find the first '{' before "decision"
            dec_pos = text.find('"decision"')
            if dec_pos > 0:
                brace_start = text.rfind('{', 0, dec_pos)
                if brace_start >= 0:
                    # Use bracket counter to find matching '}'
                    depth = 1
                    pos = brace_start + 1
                    while pos < len(text) and depth > 0:
                        if text[pos] == '{':
                            depth += 1
                        elif text[pos] == '}':
                            depth -= 1
                        pos += 1
                    if depth == 0:
                        json_str = text[brace_start:pos]
                        result = json.loads(json_str)
                        decision = _normalize_decision(str(result.get("decision", "Exclude")))
                        reason = str(result.get("reason", "")).strip()
                        # Enrich with step details if available
                        for key in ("step1_doc_type", "step2_population", "step3_tool", "step4_design"):
                            val = result.get(key, "")
                            if val and not reason.startswith(f"[{key}]"):
                                reason = f"[{key}] {val} | " + reason
                        return decision, reason
        except (json.JSONDecodeError, Exception):
            pass

        # Fallback: old flat-JSON regex
        match = re.search(r'\{[^{}]*"decision"\s*:\s*"[^"]*"[^{}]*\}', text, re.DOTALL)
        if match:
            try:
                result = json.loads(match.group(0))
                decision = _normalize_decision(str(result.get("decision", "Exclude")))
                reason = str(result.get("reason", "")).strip()
                return decision, reason
            except json.JSONDecodeError:
                pass

    return "Error", f"Failed to parse LLM response. Raw: {raw_text[:3000]}"


# ── PDF organization by screening decision ──────────────────
# 按筛选结果把 PDF 移动到对应分类子文件夹，方便后续人工整理。
CATEGORY_DIRS = {
    "Include": "include",
    "Exclude": "exclude",
    "Uncertain": "uncertain",
    "Skip": "skip",
    "Error": "error",
}


def organize_pdfs_by_decision(ws, papers_dir, decision_col, filename_col,
                              filepath_col=None) -> dict:
    """把 papers 目录下的 PDF 按 screening_decision 移动到对应子文件夹。

    子文件夹: papers/include, papers/exclude, papers/uncertain, papers/skip, papers/error

    Args:
        ws: openpyxl worksheet（最终 Excel 的活动表）
        papers_dir: PDF 所在目录
        decision_col / filename_col / filepath_col: 列号（1-based）

    Returns:
        dict: {decision: 移动成功的篇数}
    """
    moved: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        dec = str(ws.cell(row=row_idx, column=decision_col).value or "").strip()
        if dec not in CATEGORY_DIRS:
            continue
        filename = str(ws.cell(row=row_idx, column=filename_col).value or "").strip()
        if not filename:
            continue
        # 定位源文件：优先 papers_dir/filename，其次 filepath 列
        src = papers_dir / filename
        if not src.exists() and filepath_col:
            fp = str(ws.cell(row=row_idx, column=filepath_col).value or "").strip()
            if fp and Path(fp).exists():
                src = Path(fp)
        if not src.exists():
            continue
        # 目标子文件夹
        cat_dir = papers_dir / CATEGORY_DIRS[dec]
        cat_dir.mkdir(parents=True, exist_ok=True)
        dst = cat_dir / filename
        if dst.exists():
            continue  # 已归档过，跳过
        try:
            shutil.move(str(src), str(dst))
            moved[dec] = moved.get(dec, 0) + 1
        except OSError as e:
            logger.warning(f"organize: move failed {filename}: {e}")
    return moved


def run_phase4(
    inclusion_criteria: str,
    exclusion_criteria: str,
    excel_path: str = "",
    output_dir: str = "",
    mineru_output_dir: str = "",
    api_key: str = "",
    api_base: str = "https://api.deepseek.com",
    model: str = "deepseek-v4-pro",
    max_content_chars: int = 90000,
    temperature: float = 0.1,
    api_delay: float = 2.0,
    organize_pdfs: bool = True,
    progress_callback: StatusCallback = None,
) -> dict:
    """
    Run Phase 4: LLM-based full-text screening.

    Args:
        inclusion_criteria: User-defined inclusion criteria (Chinese or English).
        exclusion_criteria: User-defined exclusion criteria (Chinese or English).
        excel_path: Path to Phase 3 Excel. Auto-detect if empty.
        output_dir: Root output directory for the final screened Excel.
        mineru_output_dir: Directory where MinerU put .md files.
        api_key: DeepSeek API key.
        api_base: DeepSeek API base URL.
        model: Model name (e.g. deepseek-chat).
        max_content_chars: Max MD content chars to send per paper (default 85K for DeepSeek V3 128K context).
        temperature: LLM temperature (0.0-2.0). Lower = more deterministic. Default 0.1 for reproducible screening.
        api_delay: Seconds between API calls.
        progress_callback: Optional callback(status, data).

    Returns:
        Stats dict with counts and output path.
    """
    if not api_key:
        msg = "DeepSeek API key is required. Set DEEPSEEK_API_KEY in config.py or provide via Web UI."
        logger.error(msg)
        if progress_callback:
            progress_callback("phase4_error", {"error": msg})
        return {"error": msg}

    # Use built-in defaults only when the user supplied NO criteria at all.
    # If the user provided at least one criterion, keep their input verbatim
    # (the prompt builder will then use the custom template and ignore the
    # built-in 4-step decision tree entirely).
    if not inclusion_criteria.strip() and not exclusion_criteria.strip():
        inclusion_criteria = DEFAULT_INCLUSION_SUMMARY
        exclusion_criteria = DEFAULT_EXCLUSION_SUMMARY

    # ── Resolve paths ──
    out_root = Path(output_dir).resolve() if output_dir else BASE_DIR / "output"
    mineru_dir = Path(mineru_output_dir).resolve() if mineru_output_dir else out_root / "ocr_output"

    # Find input Excel: prefer an interrupted checkpoint (resume support),
    # otherwise fall back to the latest Phase 3 Excel.
    checkpoint_path = out_root / "results_final_screened_inprogress.xlsx"
    if not excel_path:
        if checkpoint_path.exists():
            excel_path = str(checkpoint_path)
            logger.info("Phase 4: resuming from interrupted checkpoint")
        else:
            found = _find_latest_phase3_excel(mineru_dir, out_root)
            if not found:
                msg = "No Phase 3 Excel found. Run Phase 3 (MinerU OCR) first."
                logger.error(msg)
                if progress_callback:
                    progress_callback("phase4_error", {"error": msg})
                return {"error": msg}
            excel_path = str(found)

    logger.info(f"Phase 4: source Excel = {excel_path}")
    logger.info(f"Phase 4: MinerU MD dir = {mineru_dir}")
    if progress_callback:
        progress_callback("phase4_starting", {
            "excel": excel_path,
            "mineru_dir": str(mineru_dir),
            "model": model,
        })

    # ── Load Excel ──
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[h.lower().replace(" ", "_")] = i + 1  # 1-based

    # ── Resume support: reuse pre-existing screening columns if present ──
    existing_decision_col = col_map.get("screening_decision")
    existing_reason_col = col_map.get("screening_reason")

    # Locate key columns (1-based)
    doi_col = col_map.get("doi")
    filename_col = col_map.get("filename")
    filepath_col = col_map.get("filepath")
    dl_success_col = col_map.get("download_success")
    ocr_success_col = col_map.get("ocr_success")

    missing = []
    if not doi_col:
        missing.append("DOI")
    if not filename_col:
        missing.append("Filename")
    if not dl_success_col:
        missing.append("download_success")
    if not ocr_success_col:
        missing.append("ocr_success")
    if missing:
        msg = f"Missing columns in Excel: {', '.join(missing)}. Headers: {headers}"
        logger.error(msg)
        if progress_callback:
            progress_callback("phase4_error", {"error": msg})
        return {"error": msg}

    # ── Find papers to screen ──
    # Strategy: check if MD file exists for every paper, regardless of
    # download_success / ocr_success columns (which may be stale for
    # manually-added papers).  MD existence is the ground truth.
    #
    # Matching order:
    #   1. Filename stem → exact MD index lookup (for downloaded papers)
    #   2. DOI → manual_DOI stem lookup (for manual_added / no_doi papers
    #      where OCR folder uses "manual_10.xxx" naming)
    #
    # Pre-build the MD index once for fast O(1) lookup.
    global _md_index
    if _md_index is None:
        _md_index = _build_md_index(mineru_dir)
    logger.info(f"Phase 4: MD index ready ({len(_md_index)} entries)")

    papers_to_screen = []  # (row_idx, doi, filename, md_path_or_none, dl_val)
    skipped_no_md = 0
    for row_idx in range(2, ws.max_row + 1):
        doi_raw = str(ws.cell(row=row_idx, column=doi_col).value or "").strip()
        filename = str(ws.cell(row=row_idx, column=filename_col).value or "").strip()
        dl_val = str(ws.cell(row=row_idx, column=dl_success_col).value or "").strip()

        if not doi_raw:
            continue

        md_path = None

        if filename:
            # Try exact filename stem match
            stem = Path(filename).stem
            if stem in _md_index:
                md_path = _md_index[stem]

        if md_path is None:
            # Fallback: DOI-based lookup (handles manual_10.xxx naming)
            # Step A — try with cleaned DOI (strips trailing junk like "     %(title...")
            doi_clean = _clean_doi(doi_raw)
            doi_slug = doi_clean.replace("/", "_")
            for candidate in (f"manual_{doi_slug}", doi_slug):
                if candidate in _md_index:
                    md_path = _md_index[candidate]
                    if not filename:
                        filename = f"{candidate}.pdf"
                    break

        if md_path is None:
            # Step B — case-insensitive fallback (German titles, Unicode chars)
            # Walk the index looking for a stem that starts with "manual_DOI_slug"
            # (same DOI but different case / trailing text in the OCR folder name)
            search_prefix = f"manual_{doi_slug}".lower()
            for stem_key, stem_path in _md_index.items():
                if stem_key.lower().startswith(search_prefix[:60]):
                    md_path = stem_path
                    if not filename:
                        filename = f"{stem_key}.pdf"
                    break

        if md_path is None and filepath_col:
            # Step C — filepath-based fallback (for papers whose PDF was renamed
            # but the Filename column was never updated, e.g. manual_no_doi_*)
            filepath_val = str(ws.cell(row=row_idx, column=filepath_col).value or "").strip()
            if filepath_val:
                fp_stem = Path(filepath_val).stem
                if fp_stem in _md_index:
                    md_path = _md_index[fp_stem]
                    if not filename:
                        filename = f"{fp_stem}.pdf"

        papers_to_screen.append((row_idx, doi_raw, filename, md_path, dl_val))
        if md_path is None:
            skipped_no_md += 1

    logger.info(
        f"Phase 4: {len(papers_to_screen)} papers total, "
        f"{len(papers_to_screen) - skipped_no_md} have MD files, "
        f"{skipped_no_md} skipped (no MD found)"
    )

    total = len(papers_to_screen)
    if total == 0:
        msg = "No papers found in the Excel. Check the input file."
        if progress_callback:
            progress_callback("phase4_error", {"error": msg})
        return {"error": msg}

    logger.info(f"Phase 4: {total} papers to screen")

    # ── Add (or reuse) screening columns ──
    decision_col = existing_decision_col or (len(headers) + 1)
    reason_col = existing_reason_col or (len(headers) + 2)

    header_fill = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    if existing_decision_col is None:
        for col, name in [(decision_col, "screening_decision"), (reason_col, "screening_reason")]:
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Prepare LLM client ──
    client = _get_client(api_key, api_base)

    if progress_callback:
        progress_callback("phase4_screening", {
            "total": total,
            "inclusion_criteria": "4-step decision tree (built-in)",
            "exclusion_criteria": "See decision tree steps",
        })

    # ── Screen each paper ──
    include_count = 0
    exclude_count = 0
    uncertain_count = 0
    skip_count = 0
    error_count = 0

    green_fill = PatternFill(start_color="E1F5EE", end_color="E1F5EE", fill_type="solid")
    red_fill = PatternFill(start_color="FAECE7", end_color="FAECE7", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    gray_fill = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid")

    retry_queue: list[tuple[int, str, str]] = []  # (row_idx, doi, content) for errored papers

    # ── Choose prompt template ──
    # When the user supplies their own inclusion/exclusion criteria (different
    # from the built-in defaults), use those verbatim and IGNORE the built-in
    # 4-step decision tree entirely.  Otherwise fall back to the built-in tree.
    user_has_custom = (
        inclusion_criteria.strip() != DEFAULT_INCLUSION_SUMMARY.strip()
        or exclusion_criteria.strip() != DEFAULT_EXCLUSION_SUMMARY.strip()
    )
    if user_has_custom:
        prompt_template = CUSTOM_SCREENING_USER_PROMPT
        static_kwargs = {
            "inclusion_criteria": inclusion_criteria.strip(),
            "exclusion_criteria": exclusion_criteria.strip(),
        }
        logger.info("Phase 4: using user-provided criteria (built-in tree ignored)")
    else:
        prompt_template = SCREENING_USER_PROMPT
        static_kwargs = {"extra_criteria": ""}

    resume_skipped = 0
    for i, (row_idx, doi, filename, md_path, dl_val) in enumerate(papers_to_screen):
        # ── Resume: skip papers that already have a valid decision ──
        if existing_decision_col is not None:
            prev = str(ws.cell(row=row_idx, column=existing_decision_col).value or "").strip()
            if prev in ("Include", "Exclude", "Uncertain", "Skip"):
                if prev == "Include":
                    include_count += 1
                elif prev == "Exclude":
                    exclude_count += 1
                elif prev == "Uncertain":
                    uncertain_count += 1
                else:
                    skip_count += 1
                resume_skipped += 1
                if progress_callback:
                    progress_callback("phase4_progress", {
                        "current": i + 1, "total": total,
                        "pct": int((i + 1) / total * 100),
                        "include": include_count, "exclude": exclude_count,
                        "uncertain": uncertain_count, "skip": skip_count,
                        "error": error_count,
                        "current_doi": doi[:40], "decision": prev,
                        "reason": "(resumed)",
                    })
                continue
        if md_path is None:
            if dl_val != "Yes":
                # PDF not downloaded — full text not found
                decision = "Exclude"
                reason = "未找到全文"
                exclude_count += 1
            else:
                # OCR failed but PDF was downloaded
                decision = "Skip"
                reason = "MD file not available (OCR failed)"
                skip_count += 1
        else:
            if not md_path.exists():
                decision = "Skip"
                reason = f"MD file not found: {md_path}"
                skip_count += 1
            else:
                # Read MD content
                try:
                    content = md_path.read_text(encoding="utf-8", errors="replace")
                except Exception as e:
                    logger.warning(f"Failed to read {md_path}: {e}")
                    decision = "Error"
                    reason = f"Failed to read MD file: {e}"
                    error_count += 1
                    # Write and continue
                    ws.cell(row=row_idx, column=decision_col, value=decision)
                    ws.cell(row=row_idx, column=reason_col, value=reason)
                    ws.cell(row=row_idx, column=decision_col).fill = gray_fill
                    continue

                # Truncate if too long
                if len(content) > max_content_chars:
                    content = content[:max_content_chars] + (
                        f"\n\n[... 内容过长，已截断。原文共 {len(content)} 字符 ...]"
                    )

                prompt = prompt_template.format(
                    paper_content=content,
                    **static_kwargs,
                )

                # Call LLM with system+user message structure
                decision = "Error"
                reason = ""
                try:
                    response = client.chat.completions.create(
                        model=model,
                        messages=[
                            {"role": "system", "content": SYSTEM_PROMPT},
                            {"role": "user", "content": prompt},
                        ],
                        temperature=temperature,
                        max_tokens=2500,
                    )
                    raw = response.choices[0].message.content.strip()
                    decision, reason = _parse_llm_response(raw)

                    if decision == "Include":
                        include_count += 1
                    elif decision == "Exclude":
                        exclude_count += 1
                    elif decision == "Uncertain":
                        uncertain_count += 1
                    else:
                        error_count += 1
                        retry_queue.append((row_idx, doi, content))

                    logger.info(f"  [{i+1}/{total}] {doi[:30]}... → {decision}")

                except Exception as e:
                    logger.error(f"  [{i+1}/{total}] LLM call failed for {doi}: {e}")
                    decision = "Error"
                    reason = f"API call failed: {str(e)[:200]}"
                    error_count += 1
                    retry_queue.append((row_idx, doi, content))

                # Rate limiting
                time.sleep(api_delay)

        # ── Write to Excel ──
        dec_cell = ws.cell(row=row_idx, column=decision_col, value=decision)
        rea_cell = ws.cell(row=row_idx, column=reason_col, value=reason)

        # Color coding
        if decision == "Include":
            dec_cell.fill = green_fill
        elif decision == "Exclude":
            dec_cell.fill = red_fill
        elif decision == "Uncertain":
            dec_cell.fill = yellow_fill
        else:
            dec_cell.fill = gray_fill

        # ── Progress update ──
        pct = int((i + 1) / total * 100)
        if progress_callback:
            progress_callback("phase4_progress", {
                "current": i + 1,
                "total": total,
                "pct": pct,
                "include": include_count,
                "exclude": exclude_count,
                "uncertain": uncertain_count,
                "skip": skip_count,
                "error": error_count,
                "current_doi": doi[:40],
                "decision": decision,
                "reason": reason[:120],
            })

        # ── Checkpoint save (resume-safe, every 5 papers) ──
        if (i + 1) % 5 == 0:
            wb.save(str(checkpoint_path))

    if resume_skipped:
        logger.info(f"Phase 4: resumed {resume_skipped} already-screened paper(s)")

    # ── Retry errored papers after the first full pass ──
    # DeepSeek 偶发返回空/非 JSON 响应，导致解析失败被标记 Error。
    # 第一轮跑完后对这些 Error 篇目自动重试，最多 MAX_RETRY_ROUNDS 轮。
    MAX_RETRY_ROUNDS = 2
    for retry_round in range(MAX_RETRY_ROUNDS):
        if not retry_queue:
            break
        logger.info(
            f"Phase 4: retry round {retry_round + 1}/{MAX_RETRY_ROUNDS} "
            f"for {len(retry_queue)} errored paper(s)"
        )
        if progress_callback:
            progress_callback("phase4_retry", {
                "round": retry_round + 1,
                "max_rounds": MAX_RETRY_ROUNDS,
                "count": len(retry_queue),
            })

        still_error: list[tuple[int, str, str]] = []
        for retry_idx, (row_idx, doi, content) in enumerate(retry_queue):
            prompt = prompt_template.format(
                paper_content=content,
                **static_kwargs,
            )
            decision = "Error"
            reason = ""
            try:
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": prompt},
                    ],
                    temperature=temperature,
                    max_tokens=2500,
                )
                raw = (response.choices[0].message.content or "").strip()
                decision, reason = _parse_llm_response(raw)
            except Exception as e:
                decision = "Error"
                reason = f"API call failed: {str(e)[:200]}"

            if decision == "Error":
                still_error.append((row_idx, doi, content))
                if progress_callback:
                    progress_callback("phase4_retry_progress", {
                        "round": retry_round + 1,
                        "max_rounds": MAX_RETRY_ROUNDS,
                        "current": retry_idx + 1,
                        "total": len(retry_queue),
                        "doi": doi,
                        "decision": "Error",
                        "reason": reason,
                    })
                time.sleep(api_delay)
                continue

            # 重试成功：更新计数 + Excel 单元格
            error_count -= 1
            if decision == "Include":
                include_count += 1
                fill = green_fill
            elif decision == "Exclude":
                exclude_count += 1
                fill = red_fill
            else:  # Uncertain
                uncertain_count += 1
                fill = yellow_fill

            dec_cell = ws.cell(row=row_idx, column=decision_col, value=decision)
            rea_cell = ws.cell(row=row_idx, column=reason_col, value=reason)
            dec_cell.fill = fill
            logger.info(f"  retry OK: {doi[:30]}... → {decision}")
            if progress_callback:
                progress_callback("phase4_retry_progress", {
                    "round": retry_round + 1,
                    "max_rounds": MAX_RETRY_ROUNDS,
                    "current": retry_idx + 1,
                    "total": len(retry_queue),
                    "doi": doi,
                    "decision": decision,
                    "reason": reason,
                })

            time.sleep(api_delay)

        retry_queue = still_error

    if retry_queue:
        logger.warning(
            f"Phase 4: {len(retry_queue)} paper(s) still Error after {MAX_RETRY_ROUNDS} retries"
        )

    # ── Adjust column widths for new columns ──
    ws.column_dimensions[get_column_letter(decision_col)].width = 18
    ws.column_dimensions[get_column_letter(reason_col)].width = 50

    # Update auto-filter to include new columns
    ws.auto_filter.ref = ws.dimensions

    # ── Save ──
    output_dir_path = out_root
    output_dir_path.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_path = output_dir_path / f"results_final_screened_{timestamp}.xlsx"
    wb.save(str(final_path))
    # Clean up checkpoint now that a complete result is saved
    try:
        if checkpoint_path.exists():
            checkpoint_path.unlink()
    except OSError:
        pass

    # ── 按筛选结果归档 PDF（默认开启；可在调用处关闭）──
    moved_pdfs: dict[str, int] = {}
    if organize_pdfs:
        papers_dir = out_root / "papers"
        if papers_dir.is_dir():
            try:
                moved_pdfs = organize_pdfs_by_decision(
                    ws, papers_dir, decision_col, filename_col, filepath_col
                )
                if moved_pdfs:
                    logger.info(f"Phase 4: PDF 归档完成 {moved_pdfs}")
                else:
                    logger.info("Phase 4: 无可归档 PDF（文件可能已归档或缺失）")
            except Exception as e:
                logger.warning(f"Phase 4: PDF 归档失败: {e}")

    stats = {
        "total_papers": total,
        "included": include_count,
        "excluded": exclude_count,
        "uncertain": uncertain_count,
        "skipped": skip_count,
        "errors": error_count,
        "pdfs_organized": moved_pdfs,
        "excel_path": str(final_path),
        "excel_name": final_path.name,
    }

    logger.info(
        f"Phase 4 complete: {include_count} include, {exclude_count} exclude, "
        f"{uncertain_count} uncertain, {skip_count} skip, {error_count} error. Excel: {final_path}"
    )

    if progress_callback:
        progress_callback("phase4_completed", stats)

    return stats
