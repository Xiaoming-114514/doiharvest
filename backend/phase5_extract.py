"""
DoiHarvest Phase 5 — 信息提取核心模块
=================================
一个自包含的流水线，对 Phase 4 筛选后的文献（或任意 PDF 文件夹）做结构化信息提取
合并为单一流程：

1. 接受一个包含 PDF 文件的文件夹路径（不依赖 CSV / DOI）。
2. 对文件夹内所有有效 PDF 运行 MinerU OCR。
3. 对每篇 OCR 完成的论文，将 Markdown 全文 + 用户自定义提示词发送给 LLM，
   并解析 JSON 响应。
4. 生成最终 Excel：每行一篇论文，列为所有论文 JSON 键的并集 + 元数据列。

关键设计：
  - 自包含：不依赖任何流水线状态或 Phase 1/2 产物。
  - 动态列：Excel 列由 LLM 输出决定，不同提示词产生不同列。
  - 断点续跑：每 5 篇论文保存一次检查点，中断后可恢复。
  - 损坏 PDF 预扫描：在 MinerU 之前隔离截断/损坏的 PDF。
  - 进度回调：通过 WebSocket 实时推送。
"""

import json
import logging
import re
import shutil
import subprocess
import time
from datetime import datetime
from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
StatusCallback = Callable[[str, dict], None] | None

# ── LLM client cache ──
_client = None
_client_config = {}


def _get_client(api_key: str, base_url: str = "https://api.deepseek.com"):
    """获取或创建缓存的 OpenAI 兼容客户端。"""
    global _client, _client_config
    new_cfg = {"api_key": api_key, "base_url": base_url}
    if _client is not None and _client_config == new_cfg:
        return _client
    from openai import OpenAI
    _client = OpenAI(api_key=api_key, base_url=base_url, timeout=120)
    _client_config = new_cfg
    return _client


# ── Prompt 模板 ─────────────────────────────────

EXTRACTION_SYSTEM_PROMPT = """\
You are an expert research assistant specializing in extracting structured information from academic papers.

Your task: read the full text of a research paper (extracted from PDF via OCR) and extract the specific information requested by the user.

CRITICAL PRINCIPLES:
1. Base your extraction SOLELY on the content of the provided paper text. Do not use external knowledge or fabricate information.
2. If a requested field cannot be found in the paper, output "N/A" for that field.
3. Quote specific values, numbers, or terms exactly as they appear in the paper when possible.
4. For descriptive fields, summarize concisely in the same language as the paper (Chinese for Chinese papers, English for English papers).
5. Output your response as a single JSON object. Do NOT include markdown code blocks (```json), explanations, or any text outside the JSON.
6. Use clear, descriptive field names as JSON keys.
7. ANTI-HALLUCINATION: If you are uncertain about a field's value (ambiguous wording, garbled OCR, unreadable symbols), output "N/A" instead of guessing. Never output tentative or hedged values.
"""

EXTRACTION_USER_TEMPLATE = """\
## 信息提取要求

{extraction_prompt}

## 论文全文（由 MinerU 从 PDF 提取的 Markdown）

---
{paper_content}
---

## 输出格式

请输出一个 JSON 对象，每个字段对应提取要求中的一项信息。
- 不要包含 ```json 标记或其他非 JSON 文本
- 字段值如果论文中未提及，填 "N/A"
- 字段值如果无法确认（原文含糊、OCR 残缺、符号看不清），同样填 "N/A"，不要猜测、不要写"可能/大概"之类的推断
- 数值类信息请保留原始单位和精度

请输出 JSON："""


# ── PDF / OCR 辅助函数 ───────────────────────────

def _pre_scan_corrupt_pdfs(
    papers_dir: Path,
    quarantine_dir: Path,
) -> list[str]:
    """
    在调用 MinerU 之前扫描目录中的截断/损坏 PDF。

    截断的下载文件有合法的 %PDF 头，但缺少 %%EOF 结尾标记。
    MinerU 遇到此类文件会崩溃，因此我们将其移入隔离目录。
    """
    quarantine_dir.mkdir(parents=True, exist_ok=True)
    quarantined: list[str] = []

    for pdf in sorted(papers_dir.glob("*.pdf")):
        try:
            with open(pdf, "rb") as fh:
                fh.seek(0, 2)
                size = fh.tell()
                fh.seek(max(0, size - 2048))
                tail = fh.read()
            if b"%%EOF" in tail:
                continue
        except OSError:
            pass

        try:
            shutil.move(str(pdf), str(quarantine_dir / pdf.name))
            quarantined.append(pdf.name)
            logger.warning(f"Phase 5: quarantined corrupt PDF: {pdf.name}")
        except OSError as e:
            logger.error(f"Phase 5: failed to quarantine {pdf.name}: {e}")

    return quarantined


def _build_md_index(ocr_dir: Path) -> dict[str, Path]:
    """遍历 OCR 输出目录一次，构建 stem -> md_path 映射。"""
    index: dict[str, Path] = {}
    for md_path in ocr_dir.glob("*/auto/*.md"):
        index[md_path.stem] = md_path
    for md_path in ocr_dir.glob("*/*.md"):
        if md_path.stem not in index:
            index[md_path.stem] = md_path
    return index


# ── LLM 响应解析 ────────────────────────────────

def _parse_llm_json(raw_text: str) -> dict:
    """
    将 LLM 的 JSON 响应解析为 dict。

    处理：
    - 纯 JSON
    - 包裹在 ```json ... ``` 中的 JSON
    - 嵌入在周围文本中的 JSON
    - 嵌套花括号
    """
    text = raw_text.strip()

    # 去掉 markdown 代码块
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*\n?", "", text)
        text = re.sub(r"\n?\s*```\s*$", "", text)

    # 直接解析
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 尝试提取最外层 { ... } 块
    first_brace = text.find("{")
    if first_brace >= 0:
        depth = 0
        pos = first_brace
        while pos < len(text):
            if text[pos] == "{":
                depth += 1
            elif text[pos] == "}":
                depth -= 1
                if depth == 0:
                    json_str = text[first_brace:pos + 1]
                    try:
                        return json.loads(json_str)
                    except json.JSONDecodeError:
                        pass
            pos += 1

    # 兜底：返回原始文本作为单字段
    return {"extraction_error": f"Failed to parse JSON. Raw: {raw_text[:2000]}"}


# ── 双通道反幻觉机制 ─────────────────────────────
# 每篇论文用两个不同温度各提取一遍（pass1 主通道 / pass2 校验通道），
# 字段级比对：
#   - 两遍一致        -> 取较完整的一遍
#   - 两遍不一致      -> 置 N/A（宁缺勿滥）
#   - 任一遍为 N/A    -> 置 N/A（另一遍的值无法相互印证）
#   - 任一遍含不确定表达（可能/大概/possibly...）-> 置 N/A

# 归一化后表示"无信息"的取值
NA_VARIANTS = {
    "", "na", "n/a", "nan", "none", "null", "nil", "-", "--",
    "not applicable", "not available", "not reported", "not mentioned",
    "not stated", "not specified", "not provided", "not given",
    "no data", "no information", "unknown", "unavailable", "missing",
    "n.r.", "nr", "nrp", "unreported",
    "未提及", "未报告", "未说明", "未提供", "未获取", "未给出", "未描述",
    "无", "无数据", "无信息", "无相关信息", "不适用", "暂无", "无此项",
    "无法获得", "缺失", "未发现",
}

# 值以这些词开头 -> 视为"不确定/猜测"（LLM 在犹豫，宁可 NA）
UNCERTAIN_PREFIXES = (
    "possibly", "probably", "maybe", "perhaps", "likely", "presumably",
    "approximately", "roughly", "approx", "around", "circa",
    "uncertain", "unclear", "not sure", "not certain", "unsure", "unknown",
    "estimated", "estimate", "guessing",
    "可能", "大概", "大约", "约", "疑似", "估计", "推测", "猜测", "左右",
)

# 字符串比较时忽略的常见英文装饰词（不影响语义的实词/数值比较）
STOPWORDS = {
    "a", "an", "the", "of", "in", "on", "at", "to", "from", "with", "without",
    "and", "or", "is", "are", "was", "were", "be", "been", "being", "for",
    "by", "as", "between", "among", "after", "before", "during", "using",
    "used", "use", "total", "participants", "participant", "patients",
    "patient", "subjects", "subject", "sample", "samples", "n", "p", "m",
    "sd", "mean", "range", "ci", "et", "al", "no", "not", "reported",
    "report", "reports", "had", "have", "has", "vs", "versus",
}


def _normalize_text(s) -> str:
    """归一化：转小写、压缩空白、去除常见标点（用于比较）。"""
    s = str(s)
    s = re.sub(r"[\s\u3000]+", " ", s).strip().lower()
    s = re.sub(
        r"[，。；、,.;:：!?！？()（）\[\]{}「」『』<>《》\"'“”‘’\-—_/\\|~*^=+#&%$@`]",
        "", s,
    )
    return s.strip()


def _is_uncertain(v) -> bool:
    """值是否带"不确定/猜测"表达（修饰性前缀，如 '可能 12 个月'）。"""
    if isinstance(v, str):
        s = _normalize_text(v)
        if not s:
            return False
        for p in UNCERTAIN_PREFIXES:
            if s.startswith(p):
                return True
    return False


def _is_na(v) -> bool:
    """值是否表示"无信息/不确定"（最终应输出 N/A）。"""
    if v is None:
        return True
    if isinstance(v, (dict, list)):
        return len(v) == 0
    if isinstance(v, bool):
        return False
    s = _normalize_text(v)
    if not s:
        return True
    if s in NA_VARIANTS:
        return True
    return _is_uncertain(v)


def _key_tokens(s: str) -> set:
    """提取非停用词的字母/中文 token。"""
    return {t for t in re.findall(r"[a-z\u4e00-\u9fff]+", s) if t not in STOPWORDS}


def _num_tokens(s: str) -> set:
    """提取数字 token（整数/小数/百分数）。"""
    return set(re.findall(r"\d+\.?\d*%?", s))


def _values_consistent(v1, v2) -> bool:
    """
    字段级一致性判断（宽松：同值不同表述视为一致，不同值一定不一致）。

    - dict：非 NA 键集合相同，且各键值逐一一致
    - list：长度相同且逐项可匹配（顺序无关）
    - 标量：归一化相等 / 数字 token 相等或子集 / 长串包含关系
    """
    if v1 is v2:
        return True
    if isinstance(v1, dict) and isinstance(v2, dict):
        k1 = {k for k, v in v1.items() if not _is_na(v)}
        k2 = {k for k, v in v2.items() if not _is_na(v)}
        if k1 != k2:
            return False
        return all(_values_consistent(v1[k], v2[k]) for k in k1)
    if isinstance(v1, dict) or isinstance(v2, dict):
        return False
    if isinstance(v1, list) and isinstance(v2, list):
        items1 = [i for i in v1 if not _is_na(i)]
        items2 = [i for i in v2 if not _is_na(i)]
        if len(items1) != len(items2):
            return False
        return all(any(_values_consistent(a, b) for b in items2) for a in items1)
    if isinstance(v1, list) or isinstance(v2, list):
        return False
    if isinstance(v1, bool) or isinstance(v2, bool):
        return str(v1).lower() == str(v2).lower()
    s1, s2 = _normalize_text(v1), _normalize_text(v2)
    if not s1 or not s2:
        return False
    if s1 == s2:
        return True
    # 数字 token 相等或子集 -> 核心数值一致（"347" vs "a total of 347 patients"）
    d1, d2 = _num_tokens(s1), _num_tokens(s2)
    if d1 and d2 and (d1 <= d2 or d2 <= d1):
        w1, w2 = _key_tokens(s1), _key_tokens(s2)
        if (w1 == w2) or (w1 and w1 <= w2) or (w2 and w2 <= w1) or not w1 or not w2:
            return True
    # 长串包含关系（"r=0.85" vs "r=0.85 (95% CI 0.80-0.92)"）
    if len(s1) >= 5 and len(s2) >= 5 and (s1 in s2 or s2 in s1):
        return True
    return False


def _empty_for(v):
    """返回与原始类型匹配的"无信息"值（列表/字典保留类型）。"""
    if isinstance(v, list):
        return []
    if isinstance(v, dict):
        return {}
    return "N/A"


def _value_len(v) -> int:
    """值的"信息量"：JSON 序列化后长度（用于取较完整的一遍）。"""
    try:
        return len(json.dumps(v, ensure_ascii=False))
    except Exception:
        return len(str(v))


def _merge_dual_pass(r1: dict, r2: dict) -> tuple:
    """
    合并两次提取结果（字段级），返回 (merged, discrepancies)。

    discrepancies: 被置为 N/A 的字段名列表（两遍结果无法相互印证）。
    """
    keys = list(dict.fromkeys(list(r1.keys()) + list(r2.keys())))
    merged: dict = {}
    discrepancies: list[str] = []
    for k in keys:
        if k == "extraction_error":
            continue
        v1, v2 = r1.get(k, "N/A"), r2.get(k, "N/A")
        na1, na2 = _is_na(v1), _is_na(v2)
        if na1 and na2:
            merged[k] = _empty_for(v2)
            continue
        if na1 != na2:
            # 一遍有值一遍没有 -> 有值的一遍无法被印证 -> N/A
            merged[k] = _empty_for(v2 if na1 else v1)
            discrepancies.append(k)
            continue
        if _values_consistent(v1, v2):
            # 一致 -> 取信息更完整的一遍
            merged[k] = v2 if _value_len(v2) > _value_len(v1) else v1
        else:
            merged[k] = _empty_for(v2)
            discrepancies.append(k)
    return merged, discrepancies


def _call_llm(client, model: str, prompt: str, temperature: float) -> dict:
    """调用一次 LLM 并解析 JSON。失败返回 {"extraction_error": ...}。"""
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ],
            temperature=temperature,
            max_tokens=8000,
        )
        raw = (response.choices[0].message.content or "").strip()
        return _parse_llm_json(raw)
    except Exception as e:
        return {"extraction_error": f"API call failed: {str(e)[:200]}"}


# ── Excel 写入 ───────────────────────────────────

def _format_value(v) -> str:
    """
    将提取结果中的任意 Python 值格式化为 Excel 单元格可读文本。

    支持嵌套结构（提示词让 LLM 输出 instruments/evidence 等数组）：
    - dict  ->  "key1: value1；key2: value2"
    - list  ->  每项用 " || " 分隔
    - 空值  ->  ""
    """
    if v is None:
        return ""
    if isinstance(v, dict):
        parts = []
        for k, val in v.items():
            if val is None or val == "" or val == [] or val == {}:
                continue
            parts.append(f"{k}: {_format_value(val)}")
        return "；".join(parts)
    if isinstance(v, list):
        return " || ".join(_format_value(item) for item in v if item is not None)
    if isinstance(v, bool):
        return "是" if v else "否"
    return str(v)


def _write_extraction_excel(
    papers_data: list[dict],
    excel_path: Path,
) -> dict:
    """
    将提取结果写入 Excel 文件。

    每篇论文一行。列 = 所有论文 JSON 键的并集 + 元数据列。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extraction Results"

    # ── 样式 ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
    green_fill = PatternFill(start_color="E1F5EE", end_color="E1F5EE", fill_type="solid")
    red_fill = PatternFill(start_color="FAECE7", end_color="FAECE7", fill_type="solid")
    gray_fill = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid")

    # ── 收集所有唯一的提取键（保持插入顺序）──
    extraction_keys: list[str] = []
    seen_keys: set[str] = set()
    for paper in papers_data:
        extracted = paper.get("extracted_data", {})
        if isinstance(extracted, dict):
            for key in extracted:
                if key not in seen_keys:
                    seen_keys.add(key)
                    extraction_keys.append(key)

    # ── 表头：元数据 + 提取字段 ──
    meta_headers = [
        "Filename", "PDF Path", "OCR Status", "Extraction Status", "Dual-Pass Status",
    ]
    headers = meta_headers + extraction_keys
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── 数据行 ──
    total = len(papers_data)
    ocr_ok = 0
    extract_ok = 0
    extract_error = 0

    for paper in papers_data:
        ocr_status = paper.get("ocr_status", "")
        extract_status = paper.get("extraction_status", "")
        dual_pass_status = paper.get("dual_pass_status", "")
        extracted = paper.get("extracted_data", {})

        if ocr_status == "Success":
            ocr_ok += 1
        if extract_status == "Success":
            extract_ok += 1
        elif extract_status == "Error":
            extract_error += 1

        row = [
            paper.get("filename", ""),
            paper.get("pdf_path", ""),
            ocr_status,
            extract_status,
            dual_pass_status,
        ]
        for key in extraction_keys:
            val = ""
            if isinstance(extracted, dict):
                val = _format_value(extracted.get(key, ""))
            row.append(val)

        ws.append(row)

        # 根据提取状态给行着色
        row_idx = ws.max_row
        ext_cell = ws.cell(row=row_idx, column=4)  # Extraction Status 列
        if ext_cell.value == "Success":
            ext_cell.fill = green_fill
        elif ext_cell.value == "Error":
            ext_cell.fill = red_fill
        else:
            ext_cell.fill = gray_fill

        # 双通道状态列着色：Consistent -> 绿色，Discrepant -> 警告色
        dp_cell = ws.cell(row=row_idx, column=5)  # Dual-Pass Status 列
        dp_val = str(dp_cell.value or "")
        if dp_val.startswith("Consistent"):
            dp_cell.fill = green_fill
        elif dp_val.startswith("Discrepant"):
            dp_cell.fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
            )

    # ── 列宽 ──
    ws.column_dimensions["A"].width = 35  # Filename
    ws.column_dimensions["B"].width = 50  # PDF Path
    ws.column_dimensions["C"].width = 14  # OCR Status
    ws.column_dimensions["D"].width = 16  # Extraction Status
    ws.column_dimensions["E"].width = 30  # Dual-Pass Status
    for i in range(len(extraction_keys)):
        col_letter = get_column_letter(6 + i)
        ws.column_dimensions[col_letter].width = 30

    # ── 预留手动填写列（LLM 不提取，供用户后续手工填入）──
    manual_cols = ["覆盖状态", "提取深度"]
    start_col = len(headers) + 1
    for i, col_name in enumerate(manual_cols):
        col_idx = start_col + i
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── 冻结表头 + 自动筛选 ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))

    return {
        "total_papers": total,
        "ocr_success": ocr_ok,
        "phase5_success": extract_ok,
        "phase5_error": extract_error,
        "excel_path": str(excel_path),
    }


# ── 主入口 ───────────────────────────────────────

def run_phase5(
    pdf_folder: str,
    extraction_prompt: str,
    output_dir: str = "",
    mineru_exe: str = "",
    mineru_backend: str = "pipeline",
    ocr_output_dir: str = "",
    api_key: str = "",
    api_base: str = "https://api.deepseek.com",
    model: str = "deepseek-chat",
    max_content_chars: int = 90000,
    temperature: float = 0.1,
    dual_pass: bool = True,
    dual_pass_temperature: float = 0.8,
    api_delay: float = 2.0,
    progress_callback: StatusCallback = None,
    stop_event=None,
) -> dict:
    """
    运行独立的信息提取流水线。

    步骤：
      1. 扫描 PDF 文件夹中的有效 PDF，隔离损坏的。
      2. 对所有有效 PDF 运行 MinerU OCR。
      3. 对每篇 OCR 完成的论文，将 MD 内容 + 提取提示词发送给 LLM。
      4. 解析 LLM JSON 响应并写入 Excel。

    Args:
        pdf_folder: 包含 PDF 文件的文件夹路径。
        extraction_prompt: 用户定义的要提取哪些信息的提示词。
        output_dir: 结果输出根目录。默认在 PDF 文件夹的兄弟目录创建。
        mineru_exe: MinerU CLI 可执行文件路径。
        mineru_backend: MinerU 后端 ("pipeline" 或 "vlm-engine")。
        ocr_output_dir: MinerU MD 输出的显式路径。为空则使用 {output_dir}/ocr_output。
        api_key: LLM API Key。
        api_base: LLM API 基础 URL。
        model: LLM 模型名。
        max_content_chars: 每篇论文发送的最大 MD 字符数。
        temperature: LLM 温度 (0.0-2.0)。
        dual_pass: 是否启用双通道反幻觉校验（每篇用两个温度各提取一遍）。
        dual_pass_temperature: 第二遍（校验通道）的温度。
        api_delay: API 调用间隔秒数。
        progress_callback: 可选回调(status, data)。
        stop_event: 可选 threading.Event，置位后流水线尽早停止。

    Returns:
        统计 dict，包含计数和输出路径。
    """
    # ── 输入校验 ──
    pdf_dir = Path(pdf_folder).resolve()
    if not pdf_dir.exists() or not pdf_dir.is_dir():
        msg = f"PDF 文件夹不存在或不是目录: {pdf_folder}"
        logger.error(msg)
        if progress_callback:
            progress_callback("phase5_error", {"error": msg})
        return {"error": msg}

    pdf_files = sorted(pdf_dir.glob("*.pdf"))
    if not pdf_files:
        msg = f"文件夹中未找到 PDF 文件: {pdf_folder}"
        logger.error(msg)
        if progress_callback:
            progress_callback("phase5_error", {"error": msg})
        return {"error": msg}

    if not extraction_prompt.strip():
        msg = "提取提示词为空。请提供要提取哪些信息的说明。"
        logger.error(msg)
        if progress_callback:
            progress_callback("phase5_error", {"error": msg})
        return {"error": msg}

    if not api_key:
        msg = "需要 LLM API Key。请在 Web 界面或 config.py 中设置。"
        logger.error(msg)
        if progress_callback:
            progress_callback("phase5_error", {"error": msg})
        return {"error": msg}

    # ── 解析输出目录 ──
    if output_dir:
        out_root = Path(output_dir).resolve()
    else:
        out_root = pdf_dir.parent / f"{pdf_dir.name}_extraction"
    out_root.mkdir(parents=True, exist_ok=True)

    # ── OCR 输出目录 ──
    if ocr_output_dir:
        ocr_dir = Path(ocr_output_dir).resolve()
    else:
        ocr_dir = out_root / "ocr_output"
    ocr_dir.mkdir(parents=True, exist_ok=True)

    # ── 检查点路径（断点续跑）──
    checkpoint_path = out_root / "extraction_inprogress.xlsx"

    total_pdfs = len(pdf_files)
    logger.info(
        f"Phase 5: {total_pdfs} PDFs in {pdf_dir}, "
        f"OCR output -> {ocr_dir}, results -> {out_root}"
    )
    if progress_callback:
        progress_callback("phase5_starting", {
            "pdf_folder": str(pdf_dir),
            "total_pdfs": total_pdfs,
            "ocr_output": str(ocr_dir),
            "output_dir": str(out_root),
        })

    # ═══════════════════════════════════════════════════
    # 步骤 1: 预扫描并隔离损坏 PDF
    # ═══════════════════════════════════════════════════
    quarantine_dir = out_root / "papers_corrupt"
    quarantined = _pre_scan_corrupt_pdfs(pdf_dir, quarantine_dir)
    if quarantined:
        logger.warning(
            f"Phase 5: 隔离了 {len(quarantined)} 个损坏 PDF -> {quarantine_dir}"
        )
        if progress_callback:
            progress_callback("phase5_warning", {
                "message": f"隔离了 {len(quarantined)} 个损坏 PDF（已移至 papers_corrupt/）",
            })

    # 隔离后重新列出 PDF
    pdf_files = sorted(pdf_dir.glob("*.pdf"))
    if not pdf_files:
        msg = "所有 PDF 均已损坏，没有可处理的文件。"
        if progress_callback:
            progress_callback("phase5_error", {"error": msg})
        return {"error": msg}

    # ═══════════════════════════════════════════════════
    # 步骤 2: 运行 MinerU OCR
    # ═══════════════════════════════════════════════════
    if not mineru_exe or not Path(mineru_exe).exists():
        msg = (
            f"MinerU 可执行文件不存在: {mineru_exe}\n"
            "请在 Web 界面或 config.py 中设置 MINERU_EXECUTABLE。"
        )
        logger.error(msg)
        if progress_callback:
            progress_callback("phase5_error", {"error": msg})
        return {"error": msg}

    cmd = [
        mineru_exe,
        "-p", str(pdf_dir),
        "-o", str(ocr_dir),
        "-b", mineru_backend,
    ]

    logger.info(f"Phase 5: 运行 MinerU: {' '.join(cmd)}")
    if progress_callback:
        progress_callback("phase5_ocr_running", {
            "command": " ".join(cmd),
            "total_pdfs": len(pdf_files),
        })

    try:
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            cwd=str(Path(mineru_exe).parent.parent),
        )

        line_count = 0
        for line in process.stdout:
            line = line.rstrip()
            if not line:
                continue
            line_count += 1
            logger.info(f"[MinerU] {line}")
            if progress_callback:
                progress_callback("phase5_ocr_log", {
                    "line": line[:200],
                    "line_count": line_count,
                })
            if stop_event is not None and stop_event.is_set():
                process.terminate()
                break

        process.wait()
        return_code = process.returncode

        if return_code != 0:
            logger.warning(f"MinerU exited with code {return_code} (output may still be valid)")
            if progress_callback:
                progress_callback("phase5_warning", {
                    "message": f"MinerU exited with code {return_code} (output may still be valid)",
                })
        else:
            logger.info("MinerU completed successfully")
            if progress_callback:
                progress_callback("phase5_ocr_done", {
                    "message": "MinerU OCR 处理完成",
                })

    except Exception as e:
        msg = f"Failed to run MinerU: {e}"
        logger.error(msg)
        if progress_callback:
            progress_callback("phase5_error", {"error": msg})
        return {"error": msg}

    # ═══════════════════════════════════════════════════
    # 步骤 3: 构建 MD 索引并准备提取
    # ═══════════════════════════════════════════════════
    md_index = _build_md_index(ocr_dir)
    logger.info(f"Phase 5: MD index built ({len(md_index)} entries)")

    papers_data: list[dict] = []
    for pdf in pdf_files:
        stem = pdf.stem
        md_path = md_index.get(stem)
        if md_path is None:
            # 尝试大小写不敏感的前缀匹配
            for idx_stem, idx_path in md_index.items():
                if idx_stem.lower() == stem.lower():
                    md_path = idx_path
                    break
        if md_path is None:
            # 尝试文件夹级搜索
            folder = ocr_dir / stem
            if folder.is_dir():
                auto_mds = list(folder.glob("auto/*.md"))
                if auto_mds:
                    md_path = auto_mds[0]
                else:
                    top_mds = list(folder.glob("*.md"))
                    if top_mds:
                        md_path = top_mds[0]

        ocr_status = "Success" if md_path and md_path.exists() else "Failed"
        papers_data.append({
            "filename": pdf.name,
            "pdf_path": str(pdf),
            "md_path": str(md_path) if md_path else "",
            "ocr_status": ocr_status,
            "extraction_status": "Pending",
            "extracted_data": {},
        })

    ocr_success_count = sum(1 for p in papers_data if p["ocr_status"] == "Success")
    ocr_failed_count = total_pdfs - ocr_success_count

    logger.info(
        f"Phase 5: OCR results — {ocr_success_count} success, "
        f"{ocr_failed_count} failed"
    )
    if progress_callback:
        progress_callback("phase5_ocr_summary", {
            "ocr_success": ocr_success_count,
            "ocr_failed": ocr_failed_count,
            "total": total_pdfs,
        })

    # ═══════════════════════════════════════════════════
    # 步骤 4: 断点续跑 — 从检查点加载已处理论文
    # ═══════════════════════════════════════════════════
    resume_skipped = 0
    if checkpoint_path.exists():
        try:
            old_wb = openpyxl.load_workbook(str(checkpoint_path))
            old_ws = old_wb.active
            old_headers = [cell.value for cell in old_ws[1]]
            ext_col_idx = None
            dp_col_idx = None
            for i, h in enumerate(old_headers):
                h_norm = str(h).lower().replace(" ", "_").replace("-", "_")
                if h and "extraction_status" in h_norm:
                    ext_col_idx = i + 1
                if h and "dual_pass_status" in h_norm:
                    dp_col_idx = i + 1
            # 提取字段起始列：第 5 列起，跳过 Dual-Pass Status 列
            if ext_col_idx:
                for row_idx in range(2, old_ws.max_row + 1):
                    fn_val = str(old_ws.cell(row=row_idx, column=1).value or "").strip()
                    ext_val = str(old_ws.cell(row=row_idx, column=ext_col_idx).value or "").strip()
                    if not fn_val:
                        continue
                    for paper in papers_data:
                        if paper["filename"] == fn_val and ext_val in ("Success", "Error", "Skipped"):
                            paper["extraction_status"] = ext_val
                            if dp_col_idx:
                                paper["dual_pass_status"] = str(
                                    old_ws.cell(row=row_idx, column=dp_col_idx).value or ""
                                )
                            extracted = {}
                            for col_i in range(5, len(old_headers) + 1):
                                if dp_col_idx and col_i == dp_col_idx:
                                    continue  # 跳过 Dual-Pass Status 列
                                key = old_headers[col_i - 1]
                                if key and key not in ("Filename", "PDF Path", "OCR Status", "Extraction Status", "Dual-Pass Status"):
                                    val = old_ws.cell(row=row_idx, column=col_i).value
                                    if val is not None:
                                        extracted[str(key)] = str(val)
                            paper["extracted_data"] = extracted
                            resume_skipped += 1
                            break
            logger.info(f"Phase 5: 续跑 {resume_skipped} 篇已处理论文")
        except Exception as e:
            logger.warning(f"Phase 5: 加载检查点失败: {e}")

    # ═══════════════════════════════════════════════════
    # 步骤 5: 逐篇进行 LLM 提取
    # ═══════════════════════════════════════════════════
    client = _get_client(api_key, api_base)
    papers_to_process = [p for p in papers_data if p["extraction_status"] == "Pending"]

    if progress_callback:
        progress_callback("phase5_llm_starting", {
            "total": len(papers_data),
            "to_process": len(papers_to_process),
            "model": model,
            "dual_pass": dual_pass,
            "dual_pass_temperature": dual_pass_temperature,
        })

    extract_ok = 0
    extract_error = 0
    extract_skip = 0

    # 统计已续跑的成功/失败/跳过
    for p in papers_data:
        if p["extraction_status"] == "Success":
            extract_ok += 1
        elif p["extraction_status"] == "Error":
            extract_error += 1
        elif p["extraction_status"] == "Skipped":
            extract_skip += 1

    retry_queue: list[dict] = []  # 待重试的论文

    for i, paper in enumerate(papers_to_process):
        if stop_event is not None and stop_event.is_set():
            logger.info("Phase 5: 用户请求停止，正在保存当前进度 ...")
            if progress_callback:
                progress_callback("phase5_stopping", {
                    "message": "用户请求停止，正在保存当前进度 ...",
                })
            break

        md_path_str = paper["md_path"]
        filename = paper["filename"]

        if not md_path_str or not Path(md_path_str).exists():
            paper["extraction_status"] = "Skipped"
            paper["extracted_data"] = {"error": "OCR output not found"}
            extract_skip += 1
            if progress_callback:
                progress_callback("phase5_progress", {
                    "current": i + 1,
                    "total": len(papers_to_process),
                    "pct": int((i + 1) / len(papers_to_process) * 100),
                    "current_file": filename[:50],
                    "status": "Skipped",
                    "phase5_success": extract_ok,
                    "phase5_error": extract_error,
                    "phase5_skip": extract_skip,
                })
            continue

        # 读取 MD 内容
        try:
            content = Path(md_path_str).read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            logger.warning(f"Failed to read {md_path_str}: {e}")
            paper["extraction_status"] = "Error"
            paper["extracted_data"] = {"error": f"Failed to read MD: {e}"}
            extract_error += 1
            retry_queue.append(paper)
            continue

        # 过长时截断
        if len(content) > max_content_chars:
            content = content[:max_content_chars] + (
                f"\n\n[... 内容过长，已截断。原文共 {len(content)} 字符 ...]"
            )

        prompt = EXTRACTION_USER_TEMPLATE.format(
            extraction_prompt=extraction_prompt.strip(),
            paper_content=content,
        )

        # ═══ 双通道反幻觉提取：pass1 主通道 + pass2 校验通道 ═══
        parsed1 = _call_llm(client, model, prompt, temperature)
        time.sleep(api_delay)
        pass1_ok = "extraction_error" not in parsed1
        pass2_ran = False

        if not pass1_ok:
            paper["extraction_status"] = "Error"
            paper["extracted_data"] = parsed1
            paper["dual_pass_status"] = ""
            extract_error += 1
            retry_queue.append(paper)
            logger.error(
                f"  [{i+1}/{len(papers_to_process)}] LLM pass1 失败 {filename}: "
                f"{parsed1.get('extraction_error', '')[:120]}"
            )
        else:
            if dual_pass:
                # pass1 完成，进入校验通道（仅更新 UI 进度，不写日志）
                if progress_callback:
                    progress_callback("phase5_progress", {
                        "current": i + 1,
                        "total": len(papers_to_process),
                        "pct": int((i + 1) / len(papers_to_process) * 100),
                        "current_file": filename[:50],
                        "status": "Running",
                        "pass": 1,
                        "dual_pass": True,
                        "phase5_success": extract_ok,
                        "phase5_error": extract_error,
                        "phase5_skip": extract_skip,
                    })
                parsed2 = _call_llm(client, model, prompt, dual_pass_temperature)
                time.sleep(api_delay)
                pass2_ran = True

                if "extraction_error" in parsed2:
                    paper["extraction_status"] = "Error"
                    paper["extracted_data"] = parsed2
                    paper["dual_pass_status"] = ""
                    extract_error += 1
                    retry_queue.append(paper)
                    logger.error(
                        f"  [{i+1}/{len(papers_to_process)}] LLM pass2 失败 {filename}: "
                        f"{parsed2.get('extraction_error', '')[:120]}"
                    )
                else:
                    parsed, discrepancies = _merge_dual_pass(parsed1, parsed2)
                    paper["extraction_status"] = "Success"
                    paper["extracted_data"] = parsed
                    paper["dual_pass_status"] = (
                        "Discrepant: " + "; ".join(discrepancies)
                        if discrepancies else "Consistent"
                    )
                    extract_ok += 1
                    if discrepancies:
                        logger.warning(
                            f"  [{i+1}/{len(papers_to_process)}] {filename[:40]} -> Success "
                            f"（{len(discrepancies)} 个字段两遍不一致，已置 N/A: "
                            f"{'; '.join(discrepancies)}）"
                        )
                    else:
                        logger.info(
                            f"  [{i+1}/{len(papers_to_process)}] {filename[:40]} -> Success（双通道一致）"
                        )
            else:
                paper["extraction_status"] = "Success"
                paper["extracted_data"] = parsed1
                paper["dual_pass_status"] = ""
                extract_ok += 1
                logger.info(f"  [{i+1}/{len(papers_to_process)}] {filename[:40]} -> Success")

        if progress_callback:
            final_pass = 0
            if dual_pass:
                final_pass = 2 if pass2_ran else 1
            progress_callback("phase5_progress", {
                "current": i + 1,
                "total": len(papers_to_process),
                "pct": int((i + 1) / len(papers_to_process) * 100),
                "current_file": filename[:50],
                "status": paper["extraction_status"],
                "pass": final_pass,
                "dual_pass": dual_pass,
                "dual_pass_status": paper.get("dual_pass_status", ""),
                "phase5_success": extract_ok,
                "phase5_error": extract_error,
                "phase5_skip": extract_skip,
            })

        # 检查点保存（每 5 篇）
        if (i + 1) % 5 == 0:
            try:
                _write_extraction_excel(papers_data, checkpoint_path)
                logger.info(f"Phase 5: 检查点已保存 ({i+1} 篇)")
            except Exception as e:
                logger.warning(f"Phase 5: 检查点保存失败: {e}")

    # ═══════════════════════════════════════════════════
    # 步骤 6: 重试失败的论文（最多 2 轮）
    # ═══════════════════════════════════════════════════
    MAX_RETRY_ROUNDS = 2
    for retry_round in range(MAX_RETRY_ROUNDS):
        if not retry_queue or (stop_event is not None and stop_event.is_set()):
            break
        logger.info(
            f"Phase 5: 重试第 {retry_round + 1}/{MAX_RETRY_ROUNDS} 轮，"
            f"共 {len(retry_queue)} 篇失败论文"
        )
        if progress_callback:
            progress_callback("phase5_retry", {
                "round": retry_round + 1,
                "max_rounds": MAX_RETRY_ROUNDS,
                "count": len(retry_queue),
            })

        still_error: list[dict] = []
        for retry_idx, paper in enumerate(retry_queue):
            md_path_str = paper["md_path"]
            if not md_path_str or not Path(md_path_str).exists():
                still_error.append(paper)
                continue
            try:
                content = Path(md_path_str).read_text(encoding="utf-8", errors="replace")
                if len(content) > max_content_chars:
                    content = content[:max_content_chars] + (
                        f"\n\n[... 内容过长，已截断。原文共 {len(content)} 字符 ...]"
                    )
            except Exception:
                still_error.append(paper)
                continue

            prompt = EXTRACTION_USER_TEMPLATE.format(
                extraction_prompt=extraction_prompt.strip(),
                paper_content=content,
            )
            parsed1 = _call_llm(client, model, prompt, temperature)
            time.sleep(api_delay)
            dp_status = ""
            if "extraction_error" in parsed1:
                parsed = parsed1
            elif dual_pass:
                parsed2 = _call_llm(client, model, prompt, dual_pass_temperature)
                time.sleep(api_delay)
                if "extraction_error" in parsed2:
                    parsed = parsed2
                else:
                    parsed, discrepancies = _merge_dual_pass(parsed1, parsed2)
                    dp_status = (
                        "Discrepant: " + "; ".join(discrepancies)
                        if discrepancies else "Consistent"
                    )
            else:
                parsed = parsed1

            if "extraction_error" in parsed:
                paper["extraction_status"] = "Error"
                paper["extracted_data"] = parsed
                paper["dual_pass_status"] = ""
                still_error.append(paper)
                if progress_callback:
                    progress_callback("phase5_retry_progress", {
                        "round": retry_round + 1,
                        "max_rounds": MAX_RETRY_ROUNDS,
                        "current": retry_idx + 1,
                        "total": len(retry_queue),
                        "file": paper["filename"],
                        "status": "Error",
                    })
                time.sleep(api_delay)
                continue

            # 重试成功
            extract_error -= 1
            extract_ok += 1
            paper["extraction_status"] = "Success"
            paper["extracted_data"] = parsed
            paper["dual_pass_status"] = dp_status
            if dual_pass and dp_status.startswith("Discrepant"):
                logger.warning(
                    f"  retry OK: {paper['filename'][:40]}... -> Success"
                    f"（{len(discrepancies)} 个字段两遍不一致，已置 N/A）"
                )
            else:
                logger.info(
                    f"  retry OK: {paper['filename'][:40]}... -> Success"
                    + ("（双通道一致）" if dual_pass else "")
                )
            if progress_callback:
                progress_callback("phase5_retry_progress", {
                    "round": retry_round + 1,
                    "max_rounds": MAX_RETRY_ROUNDS,
                    "current": retry_idx + 1,
                    "total": len(retry_queue),
                    "file": paper["filename"],
                    "status": "Success",
                })
            time.sleep(api_delay)

        retry_queue = still_error

    if retry_queue:
        logger.warning(
            f"Phase 5: {len(retry_queue)} 篇论文在 {MAX_RETRY_ROUNDS} 轮重试后仍失败"
        )

    # ═══════════════════════════════════════════════════
    # 步骤 7: 写入最终 Excel
    # ═══════════════════════════════════════════════════
    def _attach_dual_stats(stats: dict) -> dict:
        stats["dual_pass"] = dual_pass
        stats["dual_pass_discrepant"] = sum(
            1 for p in papers_data
            if str(p.get("dual_pass_status", "")).startswith("Discrepant")
        )
        return stats

    if stop_event is not None and stop_event.is_set():
        stats = _write_extraction_excel(papers_data, checkpoint_path)
        stats = _attach_dual_stats(stats)
        stats["stopped"] = True
        stats["message"] = "任务已停止，当前进度已保存到检查点文件。"
        if progress_callback:
            progress_callback("phase5_stopped", stats)
        return stats

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_path = out_root / f"extraction_results_{timestamp}.xlsx"
    stats = _write_extraction_excel(papers_data, final_path)
    stats = _attach_dual_stats(stats)

    # 清理检查点
    try:
        if checkpoint_path.exists():
            checkpoint_path.unlink()
    except OSError:
        pass

    logger.info(
        f"Phase 5 完成: {stats['phase5_success']} 成功, "
        f"{stats['phase5_error']} 失败. Excel: {final_path}"
    )

    if progress_callback:
        progress_callback("phase5_completed", stats)

    return stats
